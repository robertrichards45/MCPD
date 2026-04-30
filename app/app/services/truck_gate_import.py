import os
import re

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

from ..extensions import db
from ..models import TruckGateCompany, TruckGateDriver, TruckGateImportRun, TruckGateVehicle


DEFAULT_TRUCK_GATE_SOURCE = os.path.join(os.path.expanduser('~'), 'Desktop', 'truck gate data base.xlsx')
TRUCK_GATE_SOURCE_SHEET = 'DATA BASE'
TRUCK_GATE_DATA_START_ROW = 11
TRUCK_GATE_USED_COLUMNS = 9


def _normalize_text(value):
    return ' '.join(str(value or '').strip().upper().split())


def _split_trailing_state(value):
    text = _normalize_text(value)
    if not text:
        return '', ''
    match = re.match(r'^(.*?)(?:\s+([A-Z]{2}))$', text)
    if not match:
        return text, ''
    return match.group(1).strip(), match.group(2).strip()


def prepare_truck_gate_row(
    *,
    driver_name='',
    license_number='',
    license_state='',
    company_name='',
    phone_number='',
    vehicle_type='',
    plate_number='',
    plate_state='',
    make_model_color='',
    destination='',
    visit_type='',
):
    driver_name = str(driver_name or '').strip()
    license_number = str(license_number or '').strip()
    license_state = str(license_state or '').strip().upper()
    company_name = str(company_name or '').strip()
    phone_number = str(phone_number or '').strip()
    vehicle_type = str(vehicle_type or '').strip()
    plate_number = str(plate_number or '').strip()
    plate_state = str(plate_state or '').strip().upper()
    make_model_color = str(make_model_color or '').strip()
    destination = str(destination or '').strip()
    visit_type = str(visit_type or '').strip()
    return {
        'driver_name': driver_name,
        'normalized_driver_name': _normalize_text(driver_name),
        'license_number': license_number,
        'license_state': license_state,
        'company_name': company_name,
        'normalized_company_name': _normalize_text(company_name),
        'phone_number': phone_number,
        'vehicle_type': vehicle_type,
        'plate_number': plate_number,
        'plate_state': plate_state,
        'make_model_color': make_model_color,
        'destination': destination,
        'visit_type': visit_type,
    }


def parse_truck_gate_workbook(source_path=None):
    if load_workbook is None:
        raise RuntimeError('Truck Gate workbook import requires openpyxl to be installed.')
    workbook_path = source_path or DEFAULT_TRUCK_GATE_SOURCE
    workbook_path = os.path.abspath(workbook_path)
    if not os.path.exists(workbook_path):
        raise FileNotFoundError(workbook_path)

    # This workbook has a very large stored "used range" (e.g. DBF4190).
    # `read_only=True` can choke on that metadata, so load it normally and
    # only read the columns we actually map.
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    if TRUCK_GATE_SOURCE_SHEET not in workbook.sheetnames:
        raise ValueError(f'Missing sheet: {TRUCK_GATE_SOURCE_SHEET}')

    sheet = workbook[TRUCK_GATE_SOURCE_SHEET]
    headers = [sheet.cell(row=1, column=index).value for index in range(1, TRUCK_GATE_USED_COLUMNS + 1)]
    rows = []
    for row in sheet.iter_rows(
        min_row=TRUCK_GATE_DATA_START_ROW,
        max_col=TRUCK_GATE_USED_COLUMNS,
        values_only=True,
    ):
        if not any(value not in (None, '') for value in row):
            continue
        driver_name = str(row[0] or '').strip()
        if not driver_name:
            continue
        license_number, license_state = _split_trailing_state(row[1])
        plate_number, plate_state = _split_trailing_state(row[5])
        rows.append(
            prepare_truck_gate_row(
                driver_name=driver_name,
                license_number=license_number,
                license_state=license_state,
                company_name=row[2],
                phone_number=row[3],
                vehicle_type=row[4],
                plate_number=plate_number,
                plate_state=plate_state,
                make_model_color=row[6],
                destination=row[7],
                visit_type=row[8],
            )
        )
    workbook.close()
    return {
        'source_path': workbook_path,
        'source_name': os.path.basename(workbook_path),
        'sheet_name': TRUCK_GATE_SOURCE_SHEET,
        'headers': headers,
        'rows': rows,
    }


def preview_truck_gate_import(source_path=None, sample_size=20):
    parsed = parse_truck_gate_workbook(source_path)
    return {
        'source_path': parsed['source_path'],
        'source_name': parsed['source_name'],
        'sheet_name': parsed['sheet_name'],
        'headers': parsed['headers'],
        'row_count': len(parsed['rows']),
        'sample_rows': parsed['rows'][:sample_size],
    }


def _find_or_create_company(row):
    company = None
    if row['normalized_company_name']:
        company = TruckGateCompany.query.filter_by(normalized_name=row['normalized_company_name']).first()
    if company:
        company.name = row['company_name'] or company.name
        company.phone_number = row['phone_number'] or company.phone_number
        company.active = True
        return company, False

    if not row['normalized_company_name']:
        return None, False

    company = TruckGateCompany(
        name=row['company_name'] or row['normalized_company_name'].title(),
        normalized_name=row['normalized_company_name'],
        phone_number=row['phone_number'] or None,
        active=True,
    )
    db.session.add(company)
    db.session.flush()
    return company, True


def _find_or_create_driver(row, company):
    driver = None
    if row['license_number'] and row['license_state']:
        driver = TruckGateDriver.query.filter_by(
            license_number=row['license_number'],
            license_state=row['license_state'],
        ).first()
    if not driver:
        driver = TruckGateDriver.query.filter_by(
            normalized_name=row['normalized_driver_name'],
            company_id=company.id if company else None,
        ).first()

    if driver:
        driver.company_id = company.id if company else driver.company_id
        driver.full_name = row['driver_name'] or driver.full_name
        driver.normalized_name = row['normalized_driver_name'] or driver.normalized_name
        driver.phone_number = row['phone_number'] or driver.phone_number
        driver.vehicle_type = row['vehicle_type'] or driver.vehicle_type
        driver.visit_type = row['visit_type'] or driver.visit_type
        driver.destination = row['destination'] or driver.destination
        driver.active = True
        return driver, False

    driver = TruckGateDriver(
        company_id=company.id if company else None,
        full_name=row['driver_name'],
        normalized_name=row['normalized_driver_name'],
        license_number=row['license_number'] or None,
        license_state=row['license_state'] or None,
        phone_number=row['phone_number'] or None,
        vehicle_type=row['vehicle_type'] or None,
        visit_type=row['visit_type'] or None,
        destination=row['destination'] or None,
        active=True,
    )
    db.session.add(driver)
    db.session.flush()
    return driver, True


def _find_or_create_vehicle(row, company, driver):
    if not row['plate_number']:
        return None, False

    vehicle = TruckGateVehicle.query.filter_by(
        plate_number=row['plate_number'],
        plate_state=row['plate_state'] or None,
    ).first()

    if vehicle:
        vehicle.company_id = company.id if company else vehicle.company_id
        vehicle.driver_id = driver.id if driver else vehicle.driver_id
        vehicle.make_model_color = row['make_model_color'] or vehicle.make_model_color
        vehicle.active = True
        return vehicle, False

    vehicle = TruckGateVehicle(
        company_id=company.id if company else None,
        driver_id=driver.id if driver else None,
        plate_number=row['plate_number'],
        plate_state=row['plate_state'] or None,
        make_model_color=row['make_model_color'] or None,
        active=True,
    )
    db.session.add(vehicle)
    db.session.flush()
    return vehicle, True


def commit_truck_gate_import(source_path=None, actor_id=None):
    parsed = parse_truck_gate_workbook(source_path)
    inserted_count = 0
    updated_count = 0

    for row in parsed['rows']:
        company, company_created = _find_or_create_company(row)
        driver, driver_created = _find_or_create_driver(row, company)
        _, vehicle_created = _find_or_create_vehicle(row, company, driver)

        inserted_count += int(company_created) + int(driver_created) + int(vehicle_created)
        updated_count += int(not company_created and company is not None)
        updated_count += int(not driver_created and driver is not None)
        updated_count += int(not vehicle_created and row['plate_number'])

    import_run = TruckGateImportRun(
        source_name=parsed['source_name'],
        source_path=parsed['source_path'],
        row_count=len(parsed['rows']),
        inserted_count=inserted_count,
        updated_count=updated_count,
        uploaded_by=actor_id,
    )
    db.session.add(import_run)
    db.session.commit()

    return {
        'source_name': parsed['source_name'],
        'source_path': parsed['source_path'],
        'row_count': len(parsed['rows']),
        'inserted_count': inserted_count,
        'updated_count': updated_count,
    }


def upsert_truck_gate_row(row):
    company, company_created = _find_or_create_company(row)
    driver, driver_created = _find_or_create_driver(row, company)
    vehicle, vehicle_created = _find_or_create_vehicle(row, company, driver)
    return {
        'company': company,
        'driver': driver,
        'vehicle': vehicle,
        'company_created': company_created,
        'driver_created': driver_created,
        'vehicle_created': vehicle_created,
    }
