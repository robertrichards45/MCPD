from __future__ import annotations

import csv
from dataclasses import dataclass
import io
import json
from pathlib import Path
import re
import xml.etree.ElementTree as ET
from typing import Iterable
from urllib.request import urlopen

from .ai_client import ask_openai

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

try:
    import docx  # type: ignore
except Exception:  # pragma: no cover
    docx = None

try:
    from pypdf import PdfReader  # type: ignore
except Exception:  # pragma: no cover
    PdfReader = None


SUPPORTED_EXTENSIONS = {'.json', '.csv', '.xlsx', '.xlsm', '.txt', '.md', '.pdf', '.docx', '.xml', '.html', '.htm'}
OLRC_USCODE_ABOUT_URL = 'https://uscode.house.gov/about_code.xhtml'
GOVINFO_USCODE_HELP_URL = 'https://www.govinfo.gov/help/uscode'


@dataclass(frozen=True)
class IngestionResult:
    source: str
    files_processed: int
    entries_built: int
    warnings: tuple[str, ...]
    entries: tuple[dict, ...]


def _clean_text(value: str) -> str:
    return ' '.join((value or '').split()).strip()


def _split_multi(value: str | list | tuple | None) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        parts = re.split(r'[|;,]', value)
        return [item.strip() for item in parts if item.strip()]
    return [str(item).strip() for item in value if str(item).strip()]


def _safe_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _derive_summary_from_text(summary: str, official_text: str, title: str) -> str:
    existing = _clean_text(summary)
    if existing:
        return existing
    text = _clean_text(official_text)
    if not text:
        return _clean_text(title)
    sentence = re.split(r'(?<=[.!?])\s+', text)[0]
    sentence = _clean_text(sentence)
    if len(sentence) > 420:
        sentence = sentence[:417].rstrip() + '...'
    return sentence or _clean_text(title)


def _derive_elements_from_text(elements: list[str], official_text: str, title: str, citation: str) -> list[str]:
    cleaned_elements = [_clean_text(item) for item in elements if _clean_text(item)]
    if cleaned_elements:
        return cleaned_elements

    text = _clean_text(official_text)
    if not text:
        fallback = _clean_text(title) or _clean_text(citation) or 'The described conduct occurred as charged.'
        return [fallback]

    # Prefer bullet-like chunks if present.
    bullet_candidates = re.split(r'(?:\s*;\s+|\s+\(\d+\)\s+|\s+\([a-z]\)\s+)', text)
    bullet_candidates = [_clean_text(item) for item in bullet_candidates if _clean_text(item)]
    if len(bullet_candidates) >= 2:
        return bullet_candidates[:6]

    first_two = re.split(r'(?<=[.!?])\s+', text)
    first_two = [_clean_text(item) for item in first_two if _clean_text(item)]
    if first_two:
        return first_two[:3]

    return [text[:320]]


def _infer_source_from_citation(citation: str, fallback_source: str) -> str:
    normalized = (citation or '').upper()
    if normalized.startswith('OCGA'):
        return 'GEORGIA'
    if normalized.startswith('ARTICLE'):
        return 'UCMJ'
    if normalized.startswith('18 USC') or normalized.startswith('USC') or ' USC ' in normalized:
        return 'FEDERAL_USC'
    if normalized.startswith('MCLBAO') or normalized.startswith('BASE ORDER'):
        return 'BASE_ORDER'
    return fallback_source


def _normalize_entry(raw: dict, fallback_source: str) -> dict | None:
    code = _clean_text(str(raw.get('code') or raw.get('citation') or ''))
    title = _clean_text(str(raw.get('title') or raw.get('offense_title') or ''))
    official_text = _clean_text(str(raw.get('official_text') or raw.get('statutory_text') or raw.get('text') or ''))
    summary = _derive_summary_from_text(
        str(raw.get('summary') or raw.get('plain_language_summary') or raw.get('description') or ''),
        official_text,
        title,
    )
    elements = _derive_elements_from_text(
        _split_multi(raw.get('elements') or raw.get('required_elements')),
        official_text,
        title,
        code,
    )
    if not (code and title and summary and elements):
        return None
    source = _infer_source_from_citation(code, str(raw.get('source') or fallback_source).upper())
    payload = {
        'offense_id': _clean_text(str(raw.get('offense_id') or code)),
        'source': source,
        'source_type': _clean_text(str(raw.get('source_type') or (
            'federal' if source == 'FEDERAL_USC' else source.lower()
        ))),
        'source_label': _clean_text(str(raw.get('source_label') or (
            'United States Code' if source == 'FEDERAL_USC' else (
                'Georgia Code' if source == 'GEORGIA' else (
                    'UCMJ / Manual for Courts-Martial' if source == 'UCMJ' else 'MCLB Albany Base Orders'
                )
            )
        ))),
        'code': code,
        'citation': code,
        'official_citation': code,
        'title_number': _clean_text(str(raw.get('title_number') or '')),
        'section_number': _clean_text(str(raw.get('section_number') or '')),
        'chapter_number': _clean_text(str(raw.get('chapter_number') or '')),
        'article_number': _clean_text(str(raw.get('article_number') or '')),
        'order_number': _clean_text(str(raw.get('order_number') or '')),
        'title': title,
        'summary': summary,
        'plain_language_summary': summary,
        'elements': elements,
        'required_elements': elements,
        'notes': _clean_text(str(raw.get('notes') or '')),
        'keywords': _split_multi(raw.get('keywords')),
        'aliases': _split_multi(raw.get('aliases')),
        'synonyms': _split_multi(raw.get('synonyms')),
        'related_codes': _split_multi(raw.get('related_codes')),
        'narrative_triggers': _split_multi(raw.get('narrative_triggers')),
        'conduct_verbs': _split_multi(raw.get('conduct_verbs')),
        'victim_context': _split_multi(raw.get('victim_context')),
        'property_context': _split_multi(raw.get('property_context')),
        'injury_context': _split_multi(raw.get('injury_context')),
        'relationship_context': _split_multi(raw.get('relationship_context')),
        'location_context': _split_multi(raw.get('location_context')),
        'federal_context': _split_multi(raw.get('federal_context')),
        'military_context': _split_multi(raw.get('military_context')),
        'traffic_context': _split_multi(raw.get('traffic_context')),
        'juvenile_context': _split_multi(raw.get('juvenile_context')),
        'drug_context': _split_multi(raw.get('drug_context')),
        'category': _clean_text(str(raw.get('category') or '')),
        'subcategory': _clean_text(str(raw.get('subcategory') or '')),
        'severity': _clean_text(str(raw.get('severity') or '')),
        'minimum_punishment': _clean_text(str(raw.get('minimum_punishment') or '')),
        'maximum_punishment': _clean_text(str(raw.get('maximum_punishment') or raw.get('max_punishment') or '')),
        'punishment_type': _clean_text(str(raw.get('punishment_type') or '')),
        'lesser_included_offenses': _split_multi(raw.get('lesser_included_offenses')),
        'alternative_offenses': _split_multi(raw.get('alternative_offenses')),
        'overlap_notes': _split_multi(raw.get('overlap_notes')),
        'officer_notes': _clean_text(str(raw.get('officer_notes') or '')),
        'jurisdiction_conditions': _split_multi(raw.get('jurisdiction_conditions')),
        'examples': _split_multi(raw.get('examples')),
        'active_flag': bool(raw.get('active_flag', True)),
        'source_version': _clean_text(str(raw.get('source_version') or '')),
        'source_reference': _clean_text(str(raw.get('source_reference') or raw.get('source_link') or '')),
        'source_reference_url': _clean_text(str(raw.get('source_reference_url') or raw.get('source_link') or '')),
        'source_file_name': _clean_text(str(raw.get('source_file_name') or '')),
        'source_page_reference': _clean_text(str(raw.get('source_page_reference') or '')),
        'official_punishment_text': _clean_text(str(raw.get('official_punishment_text') or '')),
        'official_text': official_text,
        'official_text_available': bool(raw.get('official_text_available', bool(official_text))),
        'derived_summary': _clean_text(str(raw.get('derived_summary') or '')),
        'derived_aliases': _split_multi(raw.get('derived_aliases')),
        'derived_synonyms': _split_multi(raw.get('derived_synonyms')),
        'derived_examples': _split_multi(raw.get('derived_examples')),
        'derived_triggers': _split_multi(raw.get('derived_triggers')),
        'citation_requires_verification': bool(raw.get('citation_requires_verification', False)),
        'parser_confidence': _safe_float(raw.get('parser_confidence'), 0.0),
        'enrichment_confidence': _safe_float(raw.get('enrichment_confidence'), (0.7 if raw.get('enrichment_derived') else 0.0)),
        'last_updated': _clean_text(str(raw.get('last_updated') or '')),
    }
    if source == 'FEDERAL_USC' and (not payload['title_number'] or not payload['section_number']):
        match = re.search(r'(\d+)\s*U\.?\s*S\.?\s*C\.?\s*(?:§\s*)?([0-9A-Z().\-]+)', code.upper())
        if match:
            payload['title_number'] = payload['title_number'] or match.group(1)
            payload['section_number'] = payload['section_number'] or match.group(2)
    return payload


def _read_json_payload(blob: bytes) -> list[dict]:
    parsed = json.loads(blob.decode('utf-8'))
    if isinstance(parsed, list):
        return [item for item in parsed if isinstance(item, dict)]
    if not isinstance(parsed, dict):
        return []
    if isinstance(parsed.get('entries'), list):
        return [item for item in parsed['entries'] if isinstance(item, dict)]
    merged: list[dict] = []
    for key in ('georgia_codes', 'ucmj_articles', 'base_orders', 'federal_usc_codes', 'federal_usc_entries'):
        if isinstance(parsed.get(key), list):
            merged.extend(item for item in parsed[key] if isinstance(item, dict))
    merged.extend(_expand_structured_payload(parsed))
    return merged


def _expand_structured_payload(parsed: dict) -> list[dict]:
    rows: list[dict] = []
    if not isinstance(parsed, dict):
        return rows

    titles = parsed.get('titles')
    if isinstance(titles, list):
        for title in titles:
            if not isinstance(title, dict):
                continue
            tnum = _clean_text(str(title.get('title_number') or title.get('title') or ''))
            source_hint = _clean_text(str(title.get('source') or ''))
            sections = title.get('sections') or title.get('records') or ()
            if not isinstance(sections, list):
                continue
            for sec in sections:
                if not isinstance(sec, dict):
                    continue
                section_no = _clean_text(str(sec.get('section_number') or sec.get('section') or sec.get('num') or ''))
                heading = _clean_text(str(sec.get('heading') or sec.get('title') or ''))
                text = _clean_text(str(sec.get('official_text') or sec.get('text') or sec.get('content') or ''))
                citation = _clean_text(str(sec.get('citation') or ''))
                if not citation and tnum and section_no:
                    if source_hint.lower().startswith('federal') or 'usc' in source_hint.lower():
                        citation = f'{tnum} USC {section_no}'
                    elif source_hint.lower().startswith('georgia') or re.match(r'^\d{1,2}$', tnum):
                        citation = f'OCGA {section_no}'
                if not citation:
                    continue
                rows.append(
                    {
                        'source': source_hint or '',
                        'code': citation,
                        'citation': citation,
                        'title': heading or citation,
                        'summary': _derive_summary_from_text('', text, heading or citation),
                        'elements': _derive_elements_from_text([], text, heading or citation, citation),
                        'official_text': text,
                        'official_text_available': bool(text),
                        'title_number': tnum,
                        'section_number': section_no,
                        'source_reference_url': _clean_text(str(sec.get('source_reference_url') or title.get('source_reference_url') or '')),
                        'source_version': _clean_text(str(parsed.get('source_version') or title.get('source_version') or '')),
                        'citation_requires_verification': bool(sec.get('citation_requires_verification', False)),
                    }
                )

    for key in ('articles', 'ucmj_articles'):
        articles = parsed.get(key)
        if not isinstance(articles, list):
            continue
        for article in articles:
            if not isinstance(article, dict):
                continue
            article_no = _clean_text(str(article.get('article_number') or article.get('article') or article.get('code') or ''))
            if not article_no:
                continue
            citation = article_no if article_no.lower().startswith('article') else f'Article {article_no}'
            rows.append(
                {
                    'source': article.get('source') or 'UCMJ',
                    'code': citation,
                    'citation': citation,
                    'article_number': article_no.replace('Article', '').strip(),
                    'title': _clean_text(str(article.get('title') or article.get('heading') or citation)),
                    'summary': _clean_text(str(article.get('summary') or article.get('description') or '')),
                    'elements': _split_multi(article.get('elements') or article.get('required_elements') or ''),
                    'maximum_punishment': _clean_text(str(article.get('maximum_punishment') or article.get('max_punishment') or '')),
                    'official_text': _clean_text(str(article.get('official_text') or article.get('text') or '')),
                    'official_text_available': bool(article.get('official_text') or article.get('text')),
                    'source_version': _clean_text(str(parsed.get('source_version') or article.get('source_version') or '')),
                }
            )

    orders = parsed.get('orders') or parsed.get('base_orders_structured')
    if isinstance(orders, list):
        for order in orders:
            if not isinstance(order, dict):
                continue
            order_no = _clean_text(str(order.get('order_number') or order.get('code') or order.get('citation') or ''))
            title = _clean_text(str(order.get('title') or order.get('heading') or order_no or 'MCLB Albany Order'))
            text = _clean_text(str(order.get('official_text') or order.get('text') or order.get('content') or ''))
            if not order_no:
                order_no = f"MCLBAO {title[:24].upper()}" if title else 'MCLBAO LOCAL ORDER'
            rows.append(
                {
                    'source': 'BASE_ORDER',
                    'code': order_no,
                    'citation': order_no,
                    'order_number': order_no,
                    'title': title,
                    'summary': _derive_summary_from_text('', text, title),
                    'elements': _derive_elements_from_text([], text, title, order_no),
                    'official_text': text,
                    'official_text_available': bool(text),
                    'source_file_name': _clean_text(str(order.get('source_file_name') or '')),
                    'source_page_reference': _clean_text(str(order.get('source_page_reference') or '')),
                    'source_reference_url': _clean_text(str(order.get('source_reference_url') or '')),
                    'source_version': _clean_text(str(parsed.get('source_version') or order.get('source_version') or '')),
                    'citation_requires_verification': bool(order.get('citation_requires_verification', True)),
                }
            )
    return rows


def _read_csv_payload(blob: bytes) -> list[dict]:
    text = blob.decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append({str(k or '').strip().lower(): str(v or '').strip() for k, v in row.items()})
    return rows


def _read_xlsx_payload(blob: bytes) -> list[dict]:
    if load_workbook is None:
        return []
    workbook = load_workbook(io.BytesIO(blob), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or '').strip().lower() for cell in rows[0]]
    parsed_rows: list[dict] = []
    for row in rows[1:]:
        payload = {}
        for idx, key in enumerate(headers):
            if not key:
                continue
            value = row[idx] if idx < len(row) else ''
            payload[key] = '' if value is None else str(value).strip()
        if any(payload.values()):
            parsed_rows.append(payload)
    return parsed_rows


def _extract_citations_from_text(
    text: str,
    fallback_source: str,
    source_file_name: str = '',
    source_page_reference: str = '',
) -> list[dict]:
    clean = _clean_text(text)
    if not clean:
        return []
    chunks = re.split(r'(?:(?:\n{2,})|(?:\.\s+))', clean)
    entries: list[dict] = []
    for chunk in chunks:
        sentence = _clean_text(chunk)
        if len(sentence) < 30:
            continue
        citation = ''
        source = fallback_source
        m_ocga = re.search(r'\b(?:OCGA\s+)?(\d{1,2}-\d{1,2}-\d{1,3}(?:\.\d+)?)\b', sentence, flags=re.I)
        m_article = re.search(r'\bArticle\s+(\d+[A-Za-z]?)\b', sentence, flags=re.I)
        m_base = re.search(r'\bMCLBAO\s+([A-Za-z0-9.\-\/ ]+)\b', sentence, flags=re.I)
        m_usc = re.search(r'\b(\d+\s*USC\s*\d+(?:\([a-z0-9]+\))?)\b', sentence, flags=re.I)
        if m_ocga:
            citation = f"OCGA {m_ocga.group(1)}"
            source = 'GEORGIA'
        elif m_article:
            citation = f"Article {m_article.group(1)}"
            source = 'UCMJ'
        elif m_base:
            citation = f"MCLBAO {m_base.group(1).strip()}"
            source = 'BASE_ORDER'
        elif m_usc:
            citation = re.sub(r'\s+', ' ', m_usc.group(1)).upper()
            source = 'FEDERAL_USC'
        if not citation:
            continue
        title = sentence.split(':')[0][:120]
        summary = sentence[:500]
        entries.append(
            {
                'source': source,
                'code': citation,
                'title': title,
                'summary': summary,
                'elements': [summary],
                'official_text': sentence[:2000],
                'official_text_available': True,
                'notes': 'Extracted from uploaded source text; verify final citation and elements.',
                'source_reference': 'uploaded document',
                'source_file_name': source_file_name,
                'source_page_reference': source_page_reference,
                'parser_confidence': 0.35,
            }
        )
    return entries


def _extract_base_order_rows_from_text(
    text: str,
    source_file_name: str = '',
    source_page_reference: str = '',
) -> list[dict]:
    clean = _clean_text(text)
    if not clean:
        return []
    order_match = re.search(r'\b(MCLBAO\s+[A-Za-z0-9.\-\/ ]{2,40})\b', clean, flags=re.I)
    order_number = _clean_text(order_match.group(1)) if order_match else ''
    if not order_number:
        stem = _clean_text(Path(source_file_name).stem.upper()) if source_file_name else ''
        order_number = f'MCLBAO {stem[:24]}' if stem else 'MCLBAO LOCAL ORDER'
    code = f"{order_number} {source_page_reference}".strip()
    title = clean[:120]
    summary = _derive_summary_from_text('', clean[:1500], title)
    elements = _derive_elements_from_text([], clean[:1800], title, code)
    return [
        {
            'source': 'BASE_ORDER',
            'code': code,
            'citation': order_number,
            'order_number': order_number,
            'title': title,
            'summary': summary,
            'elements': elements,
            'official_text': clean[:4000],
            'official_text_available': True,
            'source_file_name': source_file_name,
            'source_page_reference': source_page_reference,
            'notes': 'Auto-extracted from MCLB Albany base-order source text; verify order number and page reference.',
            'citation_requires_verification': not bool(order_match),
            'parser_confidence': 0.3,
        }
    ]


def _extract_federal_rows_from_text(text: str, source_reference_url: str = '') -> list[dict]:
    clean = _clean_text(text)
    if not clean:
        return []
    pattern = re.compile(r'(\d+)\s*U\.?\s*S\.?\s*C\.?\s*(?:§|sec\.?)?\s*([0-9A-Za-z().\-]+)', flags=re.I)
    rows: list[dict] = []
    for match in pattern.finditer(clean):
        title_no = match.group(1)
        section_no = match.group(2)
        code = f'{title_no} USC {section_no}'
        start = max(0, match.start() - 120)
        end = min(len(clean), match.end() + 420)
        window = clean[start:end]
        summary = _derive_summary_from_text('', window, code)
        elements = _derive_elements_from_text([], window, code, code)
        rows.append(
            {
                'source': 'FEDERAL_USC',
                'source_type': 'federal',
                'source_label': 'United States Code',
                'code': code,
                'citation': code,
                'title': code,
                'title_number': title_no,
                'section_number': section_no,
                'summary': summary,
                'plain_language_summary': summary,
                'elements': elements,
                'required_elements': elements,
                'official_text': window[:4000],
                'official_text_available': True,
                'source_reference_url': source_reference_url,
                'source_reference': source_reference_url or 'remote federal source',
                'citation_requires_verification': True,
                'parser_confidence': 0.4,
                'notes': 'Auto-parsed from remote source text; verify final citation wording.',
            }
        )
    deduped: dict[str, dict] = {}
    for row in rows:
        deduped[str(row.get('code') or '')] = row
    return list(deduped.values())


def _extract_rows_from_xml(blob: bytes, fallback_source: str) -> list[dict]:
    try:
        root = ET.fromstring(blob)
    except Exception:
        return []

    entries: list[dict] = []
    section_nodes = list(root.findall('.//{*}section')) or list(root.findall('.//section'))
    for node in section_nodes:
        num_node = node.find('.//{*}num') or node.find('.//num')
        heading_node = node.find('.//{*}heading') or node.find('.//heading')
        content_node = node.find('.//{*}content') or node.find('.//content')

        num = _clean_text(''.join(num_node.itertext()) if num_node is not None else '')
        heading = _clean_text(''.join(heading_node.itertext()) if heading_node is not None else '')
        content = _clean_text(''.join(content_node.itertext()) if content_node is not None else '')
        if not content:
            content = _clean_text(' '.join(''.join(x.itertext()) for x in list(node)[:8]))
        if not num:
            continue

        citation = num
        source = fallback_source
        if re.search(r'\b\d+\s*U\.?\s*S\.?\s*C', num, flags=re.I):
            source = 'FEDERAL_USC'
            title_sec = re.search(r'(\d+)\s*U\.?\s*S\.?\s*C\.?\s*(?:§\s*)?([0-9A-Za-z().\-]+)', num, flags=re.I)
            if title_sec:
                citation = f"{title_sec.group(1)} USC {title_sec.group(2)}"
        elif re.search(r'^\d{1,2}-\d{1,2}-\d{1,3}(?:\.\d+)?$', num):
            source = 'GEORGIA'
            citation = f'OCGA {num}'
        elif re.search(r'^article\s+\d+[a-z]?$', num, flags=re.I):
            source = 'UCMJ'
            citation = num.title()

        title = heading or citation
        summary = _derive_summary_from_text('', content, title)
        elements = _derive_elements_from_text([], content, title, citation)
        entries.append(
            {
                'source': source,
                'code': citation,
                'title': title,
                'summary': summary,
                'elements': elements,
                'official_text': content,
                'official_text_available': bool(content),
                'citation_requires_verification': True,
                'notes': 'Auto-parsed from XML source; verify citation/section mapping.',
            }
        )

    if not entries:
        text = _clean_text(' '.join(root.itertext()))
        if text:
            entries.extend(_extract_citations_from_text(text, fallback_source))
            if fallback_source in {'ALL', 'FEDERAL_USC'}:
                entries.extend(_extract_federal_rows_from_text(text))

    deduped: dict[tuple[str, str], dict] = {}
    for entry in entries:
        key = (str(entry.get('source') or ''), str(entry.get('code') or ''))
        if all(key):
            deduped[key] = entry
    return list(deduped.values())


def _read_text_from_docx(blob: bytes) -> str:
    if docx is None:
        return ''
    document = docx.Document(io.BytesIO(blob))
    lines = [paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip()]
    return '\n'.join(lines)


def _read_text_from_pdf(blob: bytes) -> str:
    if PdfReader is None:
        return ''
    reader = PdfReader(io.BytesIO(blob))
    text_parts: list[str] = []
    for page in reader.pages:
        try:
            text_parts.append(page.extract_text() or '')
        except Exception:
            continue
    return '\n'.join(text_parts).strip()


def _read_pdf_pages(blob: bytes) -> list[tuple[int, str]]:
    if PdfReader is None:
        return []
    reader = PdfReader(io.BytesIO(blob))
    pages: list[tuple[int, str]] = []
    for idx, page in enumerate(reader.pages, start=1):
        try:
            text = (page.extract_text() or '').strip()
        except Exception:
            text = ''
        pages.append((idx, text))
    return pages


def _parse_blob(filename: str, blob: bytes, fallback_source: str) -> tuple[list[dict], list[str]]:
    ext = Path(filename).suffix.lower()
    warnings: list[str] = []
    if ext == '.json':
        return _read_json_payload(blob), warnings
    if ext == '.csv':
        return _read_csv_payload(blob), warnings
    if ext in {'.xlsx', '.xlsm'}:
        rows = _read_xlsx_payload(blob)
        if not rows and load_workbook is None:
            warnings.append('openpyxl not available for xlsx parsing.')
        return rows, warnings
    if ext in {'.txt', '.md'}:
        text = blob.decode('utf-8', errors='ignore')
        entries = _extract_citations_from_text(text, fallback_source, source_file_name=filename)
        if fallback_source == 'BASE_ORDER' and not entries:
            entries.extend(_extract_base_order_rows_from_text(text, source_file_name=filename))
        return entries, warnings
    if ext in {'.html', '.htm'}:
        text = blob.decode('utf-8', errors='ignore')
        text = re.sub(r'(?is)<script.*?>.*?</script>', ' ', text)
        text = re.sub(r'(?is)<style.*?>.*?</style>', ' ', text)
        text = re.sub(r'(?is)<[^>]+>', ' ', text)
        text = _clean_text(text)
        entries = _extract_citations_from_text(text, fallback_source, source_file_name=filename)
        if fallback_source in {'ALL', 'FEDERAL_USC'}:
            entries.extend(_extract_federal_rows_from_text(text))
        if fallback_source == 'BASE_ORDER' and not entries:
            entries.extend(_extract_base_order_rows_from_text(text, source_file_name=filename))
        return entries, warnings
    if ext == '.xml':
        entries = _extract_rows_from_xml(blob, fallback_source)
        if not entries:
            warnings.append('XML parsed but no legal section rows were detected.')
        return entries, warnings
    if ext == '.pdf':
        pages = _read_pdf_pages(blob)
        if not pages:
            warnings.append('PDF text extraction unavailable; install pypdf to improve parsing.')
            return [], warnings
        entries: list[dict] = []
        for page_no, page_text in pages:
            if not page_text:
                continue
            page_ref = f'p.{page_no}'
            before = len(entries)
            entries.extend(
                _extract_citations_from_text(
                    page_text,
                    fallback_source,
                    source_file_name=filename,
                    source_page_reference=page_ref,
                )
            )
            if fallback_source == 'BASE_ORDER' and len(entries) == before:
                entries.extend(
                    _extract_base_order_rows_from_text(
                        page_text,
                        source_file_name=filename,
                        source_page_reference=page_ref,
                    )
                )
        return entries, warnings
    if ext == '.docx':
        text = _read_text_from_docx(blob)
        if not text:
            warnings.append('DOCX parsing unavailable; install python-docx to improve parsing.')
            return [], warnings
        entries = _extract_citations_from_text(text, fallback_source, source_file_name=filename)
        if fallback_source == 'BASE_ORDER' and not entries:
            entries.extend(_extract_base_order_rows_from_text(text, source_file_name=filename))
        return entries, warnings
    warnings.append(f'Unsupported file type: {ext or filename}')
    return [], warnings


def _ai_enrich_entry(entry: dict) -> dict:
    prompt = (
        "Given a legal offense record, return STRICT JSON with keys only: "
        "aliases, synonyms, narrative_triggers, conduct_verbs, category, subcategory, severity, examples. "
        "Do not invent citations or punishments. Keep lists concise and relevant.\n\n"
        f"record={json.dumps({'source': entry.get('source'), 'code': entry.get('code'), 'title': entry.get('title'), 'summary': entry.get('summary'), 'elements': entry.get('elements')})}"
    )
    response = ask_openai(prompt, None)
    if not response:
        return entry
    try:
        start = response.find('{')
        end = response.rfind('}')
        if start < 0 or end <= start:
            return entry
        payload = json.loads(response[start:end + 1])
    except Exception:
        return entry
    if isinstance(payload.get('aliases'), list):
        entry['derived_aliases'] = [str(item).strip() for item in payload['aliases'] if str(item).strip()][:20]
    if isinstance(payload.get('synonyms'), list):
        entry['derived_synonyms'] = [str(item).strip() for item in payload['synonyms'] if str(item).strip()][:20]
    if isinstance(payload.get('examples'), list):
        entry['derived_examples'] = [str(item).strip() for item in payload['examples'] if str(item).strip()][:20]
    if isinstance(payload.get('narrative_triggers'), list):
        entry['derived_triggers'] = [str(item).strip() for item in payload['narrative_triggers'] if str(item).strip()][:20]
    if isinstance(payload.get('conduct_verbs'), list):
        entry['conduct_verbs'] = [str(item).strip() for item in payload['conduct_verbs'] if str(item).strip()][:20]
    for key in ('category', 'subcategory', 'severity'):
        if payload.get(key):
            entry[key] = str(payload[key]).strip()
    entry['derived_summary'] = entry.get('derived_summary') or (str(payload.get('summary') or '').strip())
    entry['enrichment_derived'] = True
    return entry


def _compute_parser_confidence(entry: dict, ext: str = '', mode: str = '') -> float:
    base = 0.35
    if ext in {'.json', '.csv', '.xlsx', '.xlsm'}:
        base = 0.8
    elif ext in {'.xml', '.html', '.htm'}:
        base = 0.58
    elif ext in {'.pdf', '.docx', '.txt', '.md'}:
        base = 0.45
    if mode == 'url_json':
        base = max(base, 0.75)
    elif mode == 'url_text':
        base = max(base, 0.42)

    score = base
    if (entry.get('official_text') or '').strip():
        score += 0.08
    if entry.get('elements'):
        score += 0.06
    if (entry.get('aliases') or entry.get('synonyms') or entry.get('narrative_triggers')):
        score += 0.04
    if entry.get('citation_requires_verification'):
        score -= 0.12
    if str(entry.get('source') or '').upper() == 'BASE_ORDER' and not (entry.get('order_number') or '').strip():
        score -= 0.08
    return max(0.05, min(0.99, round(score, 3)))


def ingest_uploaded_files(files: Iterable, source: str = 'ALL', use_ai_enrichment: bool = False) -> IngestionResult:
    source = (source or 'ALL').upper()
    fallback_source = source if source in {'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'} else 'GEORGIA'
    parsed_entries: list[dict] = []
    warnings: list[str] = []
    file_count = 0
    for upload in files:
        if not upload or not getattr(upload, 'filename', ''):
            continue
        filename = upload.filename
        ext = Path(filename).suffix.lower()
        if ext not in SUPPORTED_EXTENSIONS:
            warnings.append(f'Skipped unsupported file: {filename}')
            continue
        file_count += 1
        blob = upload.read()
        rows, row_warnings = _parse_blob(filename, blob, fallback_source)
        warnings.extend(row_warnings)
        for row in rows:
            normalized = _normalize_entry(row, fallback_source)
            if not normalized:
                continue
            if not normalized.get('source_file_name'):
                normalized['source_file_name'] = str(filename)
            normalized['parser_confidence'] = _compute_parser_confidence(normalized, ext=ext)
            if use_ai_enrichment:
                normalized = _ai_enrich_entry(normalized)
                normalized['enrichment_confidence'] = normalized.get('enrichment_confidence') or 0.7
            parsed_entries.append(normalized)

    deduped: dict[tuple[str, str], dict] = {}
    for item in parsed_entries:
        key = (str(item.get('source') or ''), str(item.get('code') or ''))
        if not all(key):
            continue
        deduped[key] = item

    return IngestionResult(
        source=source,
        files_processed=file_count,
        entries_built=len(deduped),
        warnings=tuple(warnings[:100]),
        entries=tuple(deduped.values()),
    )


def ingest_federal_official_payload(payload: dict | list, source_reference_url: str = '') -> IngestionResult:
    rows: list[dict]
    if isinstance(payload, list):
        rows = [item for item in payload if isinstance(item, dict)]
    elif isinstance(payload, dict):
        rows = []
        for key in ('entries', 'results', 'sections', 'records', 'federal_usc_codes'):
            value = payload.get(key)
            if isinstance(value, list):
                rows.extend(item for item in value if isinstance(item, dict))
    else:
        rows = []

    built: list[dict] = []
    warnings: list[str] = []
    for row in rows:
        normalized = _normalize_entry(row, 'FEDERAL_USC')
        if not normalized:
            continue
        if source_reference_url and not normalized.get('source_reference_url'):
            normalized['source_reference_url'] = source_reference_url
        if not normalized.get('official_text'):
            normalized['official_text_available'] = False
        normalized['parser_confidence'] = _compute_parser_confidence(normalized, ext='.json', mode='url_json')
        built.append(normalized)

    deduped: dict[tuple[str, str], dict] = {}
    for item in built:
        key = (str(item.get('source') or ''), str(item.get('code') or ''))
        if all(key):
            deduped[key] = item

    if not deduped:
        warnings.append('No federal USC entries were parsed from payload.')

    return IngestionResult(
        source='FEDERAL_USC',
        files_processed=0,
        entries_built=len(deduped),
        warnings=tuple(warnings),
        entries=tuple(deduped.values()),
    )


def ingest_federal_official_from_url(url: str, timeout_sec: int = 15) -> IngestionResult:
    target = (url or '').strip()
    if not target:
        return IngestionResult(
            source='FEDERAL_USC',
            files_processed=0,
            entries_built=0,
            warnings=('No URL provided for federal ingestion.',),
            entries=(),
        )
    try:
        with urlopen(target, timeout=timeout_sec) as response:
            raw = response.read()
    except Exception as exc:
        return IngestionResult(
            source='FEDERAL_USC',
            files_processed=0,
            entries_built=0,
            warnings=(f'Failed to fetch URL: {exc}',),
            entries=(),
        )

    parsed = None
    decoded = ''
    try:
        decoded = raw.decode('utf-8', errors='ignore')
        parsed = json.loads(decoded)
    except Exception:
        parsed = None

    if parsed is None:
        extracted_rows = _extract_federal_rows_from_text(decoded, source_reference_url=target)
        if extracted_rows:
            ingest_result = ingest_federal_official_payload(extracted_rows, source_reference_url=target)
            warnings = list(ingest_result.warnings)
            warnings.append('Remote payload was not JSON; federal citations were auto-parsed from text and flagged for verification.')
            return IngestionResult(
                source='FEDERAL_USC',
                files_processed=1,
                entries_built=ingest_result.entries_built,
                warnings=tuple(warnings),
                entries=ingest_result.entries,
            )
        return IngestionResult(
            source='FEDERAL_USC',
            files_processed=1,
            entries_built=0,
            warnings=('Remote payload is not valid JSON and no federal citations were parsed from text.',),
            entries=(),
        )

    return ingest_federal_official_payload(parsed, source_reference_url=target)


def ingest_official_from_url(url: str, source: str = 'ALL', timeout_sec: int = 15, use_ai_enrichment: bool = False) -> IngestionResult:
    target = (url or '').strip()
    chosen_source = (source or 'ALL').upper()
    if chosen_source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
        chosen_source = 'ALL'
    if not target:
        return IngestionResult(
            source=chosen_source,
            files_processed=0,
            entries_built=0,
            warnings=('No URL provided.',),
            entries=(),
        )

    try:
        with urlopen(target, timeout=timeout_sec) as response:
            raw = response.read()
    except Exception as exc:
        return IngestionResult(
            source=chosen_source,
            files_processed=0,
            entries_built=0,
            warnings=(f'Failed to fetch URL: {exc}',),
            entries=(),
        )

    decoded = raw.decode('utf-8', errors='ignore')
    rows: list[dict] = []
    warnings: list[str] = []
    try:
        rows = _read_json_payload(raw)
        if not rows:
            parsed = json.loads(decoded)
            if isinstance(parsed, dict) and isinstance(parsed.get('entries'), list):
                rows = [item for item in parsed['entries'] if isinstance(item, dict)]
    except Exception:
        rows = []

    if not rows:
        rows = _extract_citations_from_text(decoded, chosen_source if chosen_source != 'ALL' else 'GEORGIA')
        if chosen_source in {'ALL', 'FEDERAL_USC'}:
            federal_rows = _extract_federal_rows_from_text(decoded, source_reference_url=target)
            if federal_rows:
                rows.extend(federal_rows)
        if not rows:
            return IngestionResult(
                source=chosen_source,
                files_processed=1,
                entries_built=0,
                warnings=('No importable legal rows found from URL payload.',),
                entries=(),
            )
        warnings.append('URL payload parsed as text extraction (non-JSON). Verify imported citations.')

    fallback = chosen_source if chosen_source in {'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'} else 'GEORGIA'
    built: list[dict] = []
    for row in rows:
        normalized = _normalize_entry(row, fallback)
        if not normalized:
            continue
        if target and not normalized.get('source_reference_url'):
            normalized['source_reference_url'] = target
        normalized['parser_confidence'] = _compute_parser_confidence(
            normalized,
            ext='.json' if not warnings else '.txt',
            mode='url_json' if not warnings else 'url_text',
        )
        if use_ai_enrichment:
            normalized = _ai_enrich_entry(normalized)
            normalized['enrichment_confidence'] = normalized.get('enrichment_confidence') or 0.7
        built.append(normalized)

    if chosen_source != 'ALL':
        built = [row for row in built if str(row.get('source') or '').upper() == chosen_source]

    deduped: dict[tuple[str, str], dict] = {}
    for item in built:
        key = (str(item.get('source') or ''), str(item.get('code') or ''))
        if all(key):
            deduped[key] = item

    if not deduped:
        warnings.append('No rows remained after normalization/filtering.')

    return IngestionResult(
        source=chosen_source,
        files_processed=1,
        entries_built=len(deduped),
        warnings=tuple(warnings),
        entries=tuple(deduped.values()),
    )


def fetch_official_uscode_source_markers(timeout_sec: int = 10) -> dict:
    markers = {
        'olrc_url': OLRC_USCODE_ABOUT_URL,
        'govinfo_url': GOVINFO_USCODE_HELP_URL,
        'olrc_reachable': False,
        'govinfo_reachable': False,
    }
    try:
        with urlopen(OLRC_USCODE_ABOUT_URL, timeout=timeout_sec) as response:
            markers['olrc_reachable'] = bool(response.status and response.status < 400)
    except Exception:
        markers['olrc_reachable'] = False
    try:
        with urlopen(GOVINFO_USCODE_HELP_URL, timeout=timeout_sec) as response:
            markers['govinfo_reachable'] = bool(response.status and response.status < 400)
    except Exception:
        markers['govinfo_reachable'] = False
    return markers
