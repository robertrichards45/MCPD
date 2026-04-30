import json
import os
import re
import zipfile
from xml.etree import ElementTree

import requests
try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None
try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None


def category_options():
    return ['General', 'Training', 'Statistics', 'CLEOC', 'Traffic', 'Reconstruction', 'Administration']


def normalize_form_family(title):
    text = (title or '').lower()
    text = re.sub(r'(?i)\b(v(?:ersion)?\s*\d+(?:\.\d+)*)\b', ' ', text)
    text = re.sub(r'(?i)\b(rev(?:ision)?\s*[a-z0-9]+)\b', ' ', text)
    text = re.sub(r'\b\d{4}[-_/]\d{2}(?:[-_/]\d{2})?\b', ' ', text)
    text = re.sub(r'[^a-z0-9]+', ' ', text)
    return ' '.join(text.split())


def display_title(filename):
    stem = os.path.splitext(os.path.basename(filename or ''))[0]
    cleaned = stem.replace('_', ' ').replace('-', ' ').strip()
    return ' '.join(cleaned.split()) or 'Uploaded Form'


def infer_category(filename):
    name = (filename or '').lower()
    rules = (
        ('Training', ('training', 'roster', 'lesson')),
        ('Statistics', ('stats', 'stat', 'worksheet', 'tracker')),
        ('CLEOC', ('cleo', 'incident', 'report', 'narrative')),
        ('Traffic', ('dd1408', '1408', 'traffic', 'citation', 'ticket')),
        ('Reconstruction', ('reconstruction', 'diagram', 'crash')),
        ('Administration', ('admin', 'policy', 'directive', 'memo')),
    )
    for category, markers in rules:
        if any(marker in name for marker in markers):
            return category
    return 'General'


def infer_version_label(filename):
    text = os.path.splitext(os.path.basename(filename or ''))[0]
    patterns = [
        r'(?i)\b(v(?:ersion)?\s*\d+(?:\.\d+)*)\b',
        r'(?i)\b(rev(?:ision)?\s*[a-z0-9]+)\b',
        r'(?i)\b(\d{4}[-_]\d{2}[-_]\d{2})\b',
        r'(?i)\b(\d{4}[-_]\d{2})\b',
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1).replace('_', '-').upper()
    return ''


def version_rank(version_label, uploaded_at=None):
    label = (version_label or '').strip()
    if not label:
        timestamp = uploaded_at.timestamp() if uploaded_at else 0
        return (0, timestamp)

    numbers = [int(part) for part in re.findall(r'\d+', label)]
    if numbers:
        padded = tuple(numbers[:6]) + tuple([0] * max(0, 6 - len(numbers[:6])))
        return (2,) + padded

    letters = re.findall(r'[A-Z]+', label.upper())
    if letters:
        score = tuple(sum((ord(char) - 64) for char in item) for item in letters[:4])
        return (1,) + score

    timestamp = uploaded_at.timestamp() if uploaded_at else 0
    return (0, timestamp)


def heuristic_metadata(filename):
    return {
        'filename': filename,
        'title': display_title(filename),
        'category': infer_category(filename),
        'version_label': infer_version_label(filename),
    }


def _infer_category_from_text(text, fallback):
    lowered = (text or '').lower()
    if not lowered:
        return fallback
    text_rules = (
        ('Training', ('training roster', 'lesson plan', 'training')),
        ('Statistics', ('statistics', 'monthly stats', 'worksheet', 'tracker')),
        ('CLEOC', ('incident report', 'narrative', 'statement', 'report')),
        ('Traffic', ('dd form 1408', 'violation notice', 'traffic ticket', 'citation')),
        ('Reconstruction', ('accident reconstruction', 'crash diagram', 'scene diagram')),
        ('Administration', ('policy', 'memorandum', 'directive', 'administrative')),
    )
    for category, markers in text_rules:
        if any(marker in lowered for marker in markers):
            return category
    return fallback


def _infer_title_from_text(text, fallback):
    lines = [line.strip() for line in (text or '').splitlines() if line.strip()]
    for line in lines[:8]:
        if len(line) >= 4:
            return ' '.join(line.split())[:200]
    return fallback


def _infer_version_from_text(text, fallback):
    lowered = text or ''
    patterns = [
        r'(?i)\b(v(?:ersion)?\s*\d+(?:\.\d+)*)\b',
        r'(?i)\b(rev(?:ision)?\s*[a-z0-9]+)\b',
        r'(?i)\b(effective\s+\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b',
        r'(?i)\b(\d{4}[-/]\d{2}[-/]\d{2})\b',
    ]
    for pattern in patterns:
        match = re.search(pattern, lowered)
        if match:
            return match.group(1).replace('_', '-').upper()
    return fallback


def inspect_uploaded_file(file_storage):
    filename = getattr(file_storage, 'filename', '') or ''
    fallback = heuristic_metadata(filename)
    snippet = ''
    ext = os.path.splitext(filename)[1].lower()
    stream = getattr(file_storage, 'stream', None)

    try:
        if ext == '.pdf' and stream is not None and PdfReader is not None:
            stream.seek(0)
            reader = PdfReader(stream)
            if reader.pages:
                snippet = (reader.pages[0].extract_text() or '')[:4000]
            stream.seek(0)
        elif ext in {'.docx', '.docm'} and stream is not None:
            stream.seek(0)
            with zipfile.ZipFile(stream) as archive:
                if 'word/document.xml' in archive.namelist():
                    xml_bytes = archive.read('word/document.xml')
                    root = ElementTree.fromstring(xml_bytes)
                    text_parts = []
                    for node in root.iter():
                        if node.tag.endswith('}t') and node.text:
                            text_parts.append(node.text)
                    snippet = ' '.join(text_parts)[:4000]
            stream.seek(0)
        elif ext in {'.xlsx', '.xlsm'} and stream is not None and load_workbook is not None:
            stream.seek(0)
            workbook = load_workbook(stream, read_only=True, data_only=True)
            text_parts = []
            for sheet in workbook.worksheets[:2]:
                for row in sheet.iter_rows(max_row=12, max_col=6, values_only=True):
                    for value in row:
                        if value is not None:
                            text_parts.append(str(value))
            snippet = ' '.join(text_parts)[:4000]
            workbook.close()
            stream.seek(0)
        elif ext in {'.html', '.htm', '.xml', '.json', '.md', '.rtf'} and stream is not None:
            stream.seek(0)
            snippet = stream.read(4000).decode('utf-8', errors='ignore')
            stream.seek(0)
        elif ext in {'.txt', '.csv'} and stream is not None:
            stream.seek(0)
            snippet = stream.read(4000).decode('utf-8', errors='ignore')
            stream.seek(0)
    except Exception:
        if stream is not None:
            try:
                stream.seek(0)
            except Exception:
                pass
        snippet = ''

    if not snippet:
        return fallback

    return {
        'filename': filename,
        'title': _infer_title_from_text(snippet, fallback['title']),
        'category': _infer_category_from_text(snippet, fallback['category']),
        'version_label': _infer_version_from_text(snippet, fallback['version_label']),
        'text_snippet': snippet[:1000],
    }


def _parse_ai_results(payload, filenames):
    if not isinstance(payload, list):
        return None
    results = []
    for index, filename in enumerate(filenames):
        item = payload[index] if index < len(payload) and isinstance(payload[index], dict) else {}
        candidate = {
            'filename': filename,
            'title': str(item.get('title') or '').strip() or display_title(filename),
            'category': str(item.get('category') or '').strip() or infer_category(filename),
            'version_label': str(item.get('version_label') or '').strip(),
        }
        if candidate['category'] not in category_options():
            candidate['category'] = infer_category(filename)
        results.append(candidate)
    return results


def detect_form_metadata_batch(filenames, api_key):
    filenames = [name for name in (filenames or []) if name]
    baseline = [heuristic_metadata(name) for name in filenames]
    if not filenames or not api_key:
        return baseline

    prompt = (
        "For each filename, infer a clean Title, Category, and Version Label for an internal police forms library. "
        "Allowed categories are: General, Training, Statistics, CLEOC, Traffic, Reconstruction, Administration. "
        "Return only valid JSON as an array of objects with keys title, category, version_label in the same order as the filenames.\n\n"
        f"Filenames: {json.dumps(filenames)}"
    )
    headers = {
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',
    }
    payload = {
        'model': 'gpt-4.1-mini',
        'input': [
            {
                'role': 'system',
                'content': 'You extract concise metadata for uploaded files. Return strict JSON only.',
            },
            {
                'role': 'user',
                'content': prompt,
            },
        ],
    }

    try:
        response = requests.post(
            'https://api.openai.com/v1/responses',
            headers=headers,
            data=json.dumps(payload),
            timeout=20,
        )
        response.raise_for_status()
        data = response.json()
        text = data.get('output', [{}])[0].get('content', [{}])[0].get('text', '')
        parsed = json.loads(text)
        ai_results = _parse_ai_results(parsed, filenames)
        return ai_results or baseline
    except Exception:
        return baseline


def detect_form_metadata_from_uploads(files, api_key):
    inspected = [inspect_uploaded_file(file) for file in (files or []) if getattr(file, 'filename', '')]
    if not inspected:
        return []

    baseline = []
    for item in inspected:
        baseline.append({
            'filename': item['filename'],
            'title': item['title'],
            'category': item['category'],
            'version_label': item['version_label'],
        })

    if not api_key:
        return baseline

    prompt_payload = [
        {
            'filename': item['filename'],
            'text_snippet': item.get('text_snippet', ''),
            'detected_title': item['title'],
            'detected_category': item['category'],
            'detected_version_label': item['version_label'],
        }
        for item in inspected
    ]
    prompt = (
        "For each uploaded file, refine the metadata for an internal police forms library. "
        "Use the filename and text snippet. Allowed categories are: General, Training, Statistics, CLEOC, Traffic, Reconstruction, Administration. "
        "Return only valid JSON as an array of objects with keys title, category, version_label in the same order.\n\n"
        f"Files: {json.dumps(prompt_payload)}"
    )
    headers = {
        'Authorization': f'Bearer {api_key}',
        'Content-Type': 'application/json',
    }
    payload = {
        'model': 'gpt-4.1-mini',
        'input': [
            {
                'role': 'system',
                'content': 'You extract concise metadata for uploaded files. Return strict JSON only.',
            },
            {
                'role': 'user',
                'content': prompt,
            },
        ],
    }

    try:
        response = requests.post(
            'https://api.openai.com/v1/responses',
            headers=headers,
            data=json.dumps(payload),
            timeout=20,
        )
        response.raise_for_status()
        data = response.json()
        text = data.get('output', [{}])[0].get('content', [{}])[0].get('text', '')
        parsed = json.loads(text)
        filenames = [item['filename'] for item in inspected]
        ai_results = _parse_ai_results(parsed, filenames)
        return ai_results or baseline
    except Exception:
        return baseline


def choose_latest_form(forms, api_key):
    candidates = [form for form in (forms or []) if form]
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]

    if api_key:
        prompt_items = [
            {
                'id': form.id,
                'title': form.title,
                'version_label': form.version_label,
                'uploaded_at': form.uploaded_at.isoformat() if getattr(form, 'uploaded_at', None) else '',
            }
            for form in candidates
        ]
        headers = {
            'Authorization': f'Bearer {api_key}',
            'Content-Type': 'application/json',
        }
        payload = {
            'model': 'gpt-4.1-mini',
            'input': [
                {
                    'role': 'system',
                    'content': 'Choose the single most current form version from the list. Return strict JSON only like {"id": 123}.',
                },
                {
                    'role': 'user',
                    'content': 'Select the most up-to-date form version.\n\n' + json.dumps(prompt_items),
                },
            ],
        }
        try:
            response = requests.post(
                'https://api.openai.com/v1/responses',
                headers=headers,
                data=json.dumps(payload),
                timeout=15,
            )
            response.raise_for_status()
            data = response.json()
            text = data.get('output', [{}])[0].get('content', [{}])[0].get('text', '')
            parsed = json.loads(text)
            selected_id = int(parsed.get('id'))
            for form in candidates:
                if form.id == selected_id:
                    return form
        except Exception:
            pass

    return max(
        candidates,
        key=lambda form: (
            version_rank(getattr(form, 'version_label', ''), getattr(form, 'uploaded_at', None)),
            getattr(form, 'uploaded_at', None) or 0,
            getattr(form, 'id', 0),
        ),
    )
