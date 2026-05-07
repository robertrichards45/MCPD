import csv
from datetime import datetime, timezone
from difflib import SequenceMatcher
import io
import json
from pathlib import Path
import re
from uuid import uuid4

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None

from flask import Blueprint, Response, abort, current_app, flash, make_response, redirect, render_template, request, send_file, session, url_for
from flask_login import current_user, login_required

from ..extensions import db
from ..services.legal_lookup import STOPWORDS, _source_file_candidates, _tokenize, LegalMatch, corpus_status, export_corpus_payload, get_entry, import_corpus_payload, reference_download_info, search_entries
from ..services.ai_client import ask_openai, is_ai_unavailable_message


bp = Blueprint('legal', __name__)
AI_HINT_CACHE_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'ai_query_hints.json'
NARRATIVE_INTERP_CACHE_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'ai_narrative_interp_cache.json'
LEGAL_QUERY_LOG_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'legal_query_log.jsonl'
LEGAL_TUNING_CASES_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'legal_tuning_cases.jsonl'
LEGAL_SEARCH_FAILURE_LOG_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'legal_search_failures.jsonl'
SOURCE_OPTIONS = ('ALL', 'STATE', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC')
LOCAL_SOURCE_OPTIONS = {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}
STATE_OPTIONS = (
    ('AL', 'Alabama'), ('AK', 'Alaska'), ('AZ', 'Arizona'), ('AR', 'Arkansas'), ('CA', 'California'),
    ('CO', 'Colorado'), ('CT', 'Connecticut'), ('DE', 'Delaware'), ('FL', 'Florida'), ('GA', 'Georgia'),
    ('HI', 'Hawaii'), ('ID', 'Idaho'), ('IL', 'Illinois'), ('IN', 'Indiana'), ('IA', 'Iowa'),
    ('KS', 'Kansas'), ('KY', 'Kentucky'), ('LA', 'Louisiana'), ('ME', 'Maine'), ('MD', 'Maryland'),
    ('MA', 'Massachusetts'), ('MI', 'Michigan'), ('MN', 'Minnesota'), ('MS', 'Mississippi'), ('MO', 'Missouri'),
    ('MT', 'Montana'), ('NE', 'Nebraska'), ('NV', 'Nevada'), ('NH', 'New Hampshire'), ('NJ', 'New Jersey'),
    ('NM', 'New Mexico'), ('NY', 'New York'), ('NC', 'North Carolina'), ('ND', 'North Dakota'), ('OH', 'Ohio'),
    ('OK', 'Oklahoma'), ('OR', 'Oregon'), ('PA', 'Pennsylvania'), ('RI', 'Rhode Island'), ('SC', 'South Carolina'),
    ('SD', 'South Dakota'), ('TN', 'Tennessee'), ('TX', 'Texas'), ('UT', 'Utah'), ('VT', 'Vermont'),
    ('VA', 'Virginia'), ('WA', 'Washington'), ('WV', 'West Virginia'), ('WI', 'Wisconsin'), ('WY', 'Wyoming'),
    ('DC', 'District of Columbia'),
)
STATE_LABELS = dict(STATE_OPTIONS)


def _no_store_response(rendered):
    response = make_response(rendered)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


def _legal_source_or_404(source: str) -> str:
    normalized = (source or 'ALL').strip().upper()
    if normalized == 'STATE':
        normalized = 'GEORGIA'
    if normalized not in {'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
        abort(404)
    return normalized


def _normalize_state(value: str) -> str:
    state = (value or 'GA').strip().upper()
    return state if state in STATE_LABELS else 'GA'


def _normalize_lookup_source(value: str, fallback='ALL') -> str:
    source = (value or fallback or 'ALL').strip().upper()
    if source in {'GEORGIA', 'GA', 'STATE_CODE'}:
        return 'STATE'
    if source in SOURCE_OPTIONS:
        return source
    if source in {'FEDERAL', 'USC', 'UNITED_STATES_CODE'}:
        return 'FEDERAL_USC'
    if source in {'MCM', 'ARTICLE'}:
        return 'UCMJ'
    if source in {'BASE', 'ORDER', 'MCLBAO'}:
        return 'BASE_ORDER'
    return fallback if fallback in SOURCE_OPTIONS else 'ALL'


def _local_source_for_state(source: str, state_code: str) -> str:
    if source == 'STATE':
        return 'GEORGIA' if state_code == 'GA' else 'STATE_UNAVAILABLE'
    return source


def _search_entries_for_scope(query: str, source: str, state_code: str, strict_gating=True) -> list[LegalMatch]:
    if not query:
        return []
    if source == 'STATE':
        return search_entries(query, 'GEORGIA', strict_gating=strict_gating) if state_code == 'GA' else []
    if source == 'ALL' and state_code != 'GA':
        merged = []
        seen = set()
        for local_source in ('UCMJ', 'BASE_ORDER', 'FEDERAL_USC'):
            for item in search_entries(query, local_source, strict_gating=strict_gating):
                key = (item.entry.source, item.entry.code)
                if key in seen:
                    continue
                merged.append(item)
                seen.add(key)
        return sorted(merged, key=lambda item: (-item.score, item.entry.source, item.entry.code))
    return search_entries(query, _local_source_for_state(source, state_code), strict_gating=strict_gating)


def _source_label_for_state(source: str, state_code: str) -> str:
    if source == 'STATE':
        return f'{STATE_LABELS.get(state_code, state_code)} State Law'
    if source == 'ALL':
        return f'{STATE_LABELS.get(state_code, state_code)} + Federal + UCMJ'
    return _source_label(source)


def _reference_text(entry) -> str:
    lines = [
        entry.code,
        entry.title,
        '',
    ]
    if entry.summary:
        lines.extend(['Summary', entry.summary, ''])
    if entry.plain_language_summary and entry.plain_language_summary != entry.summary:
        lines.extend(['Plain-Language Summary', entry.plain_language_summary, ''])
    if entry.required_elements:
        lines.append('Required Elements')
        lines.extend(f'- {item}' for item in entry.required_elements)
        lines.append('')
    elif entry.elements:
        lines.append('Elements')
        lines.extend(f'- {item}' for item in entry.elements)
        lines.append('')
    if entry.penalties or entry.minimum_punishment or entry.maximum_punishment:
        lines.append('Penalties')
        if entry.penalties:
            lines.append(entry.penalties)
        if entry.minimum_punishment:
            lines.append(f'Minimum: {entry.minimum_punishment}')
        if entry.maximum_punishment:
            lines.append(f'Maximum: {entry.maximum_punishment}')
        lines.append('')
    if entry.official_text:
        lines.extend(['Full Text', entry.official_text, ''])
    if entry.enforcement_notes:
        lines.extend(['Enforcement Notes', entry.enforcement_notes, ''])
    return '\n'.join(lines).strip() + '\n'


def _lookup_match_context(query: str, source: str, code: str):
    clean_query = (query or '').strip()
    if not clean_query:
        return None
    for item in search_entries(clean_query, source):
        if item.entry.code == code:
            return item
    return None


def _safe_ai_guidance(prompt):
    if not (prompt or '').strip():
        return ''
    answer = ask_openai(prompt, None)
    if not answer:
        return ''
    if is_ai_unavailable_message(answer):
        return ''
    return answer


def _normalize_query_key(value: str) -> str:
    cleaned = re.sub(r'[^a-z0-9\s]', ' ', (value or '').lower())
    return ' '.join(cleaned.split())


def _dedupe_phrases(values, limit=8):
    cleaned = []
    seen = set()
    for value in values or []:
        text = re.sub(r'\s+', ' ', str(value or '').strip())
        if not text:
            continue
        key = text.lower()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(text)
        if len(cleaned) >= limit:
            break
    return cleaned


def _load_ai_hint_cache() -> dict:
    try:
        if not AI_HINT_CACHE_PATH.exists():
            return {}
        data = json.loads(AI_HINT_CACHE_PATH.read_text(encoding='utf-8'))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _write_ai_hint_cache(cache: dict) -> None:
    try:
        AI_HINT_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        AI_HINT_CACHE_PATH.write_text(json.dumps(cache, indent=2), encoding='utf-8')
    except Exception:
        return


def _search_quality_bucket(results: list[LegalMatch]) -> str:
    if not results:
        return 'no_result'
    top = int(results[0].confidence)
    if top < 55:
        return 'weak'
    if top < 75:
        return 'review'
    return 'strong'


def _append_search_failure_log(query: str, source: str, results: list[LegalMatch], previous_query: str = '') -> None:
    quality = _search_quality_bucket(results)
    repeated = False
    previous = (previous_query or '').strip()
    if previous and query:
        repeated = SequenceMatcher(None, _normalize_query_key(previous), _normalize_query_key(query)).ratio() >= 0.55
    if quality == 'strong' and not repeated:
        return
    try:
        LEGAL_SEARCH_FAILURE_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            'ts_utc': datetime.now(timezone.utc).isoformat(),
            'query': query,
            'source': source,
            'quality': quality,
            'previous_query': previous if repeated else '',
            'reformulation_detected': repeated,
            'result_count': len(results),
            'top_codes': [item.entry.code for item in results[:5]],
            'top_confidence': int(results[0].confidence) if results else 0,
        }
        with LEGAL_SEARCH_FAILURE_LOG_PATH.open('a', encoding='utf-8') as handle:
            handle.write(json.dumps(payload) + '\n')
    except Exception:
        return


def _append_legal_query_log(query: str, source: str, results: list[LegalMatch]) -> None:
    if not (query or '').strip():
        return
    try:
        LEGAL_QUERY_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        quality = _search_quality_bucket(results)
        payload = {
            'ts_utc': datetime.now(timezone.utc).isoformat(),
            'query': query,
            'source': source,
            'result_count': len(results),
            'top_codes': [item.entry.code for item in results[:5]],
            'top_confidence': int(results[0].confidence) if results else 0,
            'weak_result': quality in {'no_result', 'weak'},
            'quality': quality,
        }
        with LEGAL_QUERY_LOG_PATH.open('a', encoding='utf-8') as handle:
            handle.write(json.dumps(payload) + '\n')
    except Exception:
        return


def _append_legal_tuning_case(query: str, source: str, expected_codes: list[str], notes: str = '') -> None:
    clean_query = (query or '').strip()
    if not clean_query:
        return
    try:
        LEGAL_TUNING_CASES_PATH.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            'case_id': str(uuid4()),
            'ts_utc': datetime.now(timezone.utc).isoformat(),
            'query': clean_query,
            'source': (source or 'ALL').strip().upper(),
            'expected_codes': [code.strip() for code in expected_codes if code.strip()],
            'notes': (notes or '').strip(),
        }
        with LEGAL_TUNING_CASES_PATH.open('a', encoding='utf-8') as handle:
            handle.write(json.dumps(payload) + '\n')
    except Exception:
        return


def _load_legal_tuning_cases(limit: int = 50) -> list[dict]:
    rows: list[dict] = []
    try:
        if not LEGAL_TUNING_CASES_PATH.exists():
            return rows
        with LEGAL_TUNING_CASES_PATH.open('r', encoding='utf-8') as handle:
            for raw in handle:
                line = raw.strip()
                if not line:
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if not isinstance(payload, dict):
                    continue
                query = str(payload.get('query') or '').strip()
                source = str(payload.get('source') or 'ALL').strip().upper()
                expected = payload.get('expected_codes')
                expected_codes = [str(code).strip() for code in expected] if isinstance(expected, list) else []
                if not query or not expected_codes:
                    continue
                case_id = str(payload.get('case_id') or '').strip()
                if not case_id:
                    # Backward-compatible fallback for rows saved before case_id existed.
                    case_id = f"legacy-{abs(hash(line))}"
                rows.append(
                    {
                        'case_id': case_id,
                        'ts_utc': str(payload.get('ts_utc') or '').strip(),
                        'query': query,
                        'source': source if source in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'} else 'ALL',
                        'expected_codes': expected_codes,
                        'notes': str(payload.get('notes') or '').strip(),
                    }
                )
        rows = list(reversed(rows))
        return rows[: max(1, min(limit, 500))]
    except Exception:
        return []


def _delete_legal_tuning_case(case_id: str) -> bool:
    target = (case_id or '').strip()
    if not target or not LEGAL_TUNING_CASES_PATH.exists():
        return False
    try:
        kept: list[str] = []
        removed = False
        with LEGAL_TUNING_CASES_PATH.open('r', encoding='utf-8') as handle:
            for raw in handle:
                line = raw.rstrip('\n')
                if not line.strip():
                    continue
                try:
                    payload = json.loads(line)
                except json.JSONDecodeError:
                    kept.append(line)
                    continue
                payload_case_id = str(payload.get('case_id') or '').strip()
                if payload_case_id == target:
                    removed = True
                    continue
                kept.append(line)
        if not removed:
            return False
        with LEGAL_TUNING_CASES_PATH.open('w', encoding='utf-8') as handle:
            for line in kept:
                handle.write(line + '\n')
        return True
    except Exception:
        return False


def _get_cached_ai_hint(query: str, source: str) -> dict | None:
    cache = _load_ai_hint_cache()
    key = f"{_normalize_query_key(query)}|{(source or 'ALL').upper()}"
    hit = cache.get(key)
    if not isinstance(hit, dict):
        return None
    terms = hit.get('terms') if isinstance(hit.get('terms'), list) else []
    query_variants = hit.get('query_variants') if isinstance(hit.get('query_variants'), list) else []
    related_policy_terms = hit.get('related_policy_terms') if isinstance(hit.get('related_policy_terms'), list) else []
    source_hint = str(hit.get('source_hint') or source).strip().upper()
    if source_hint == 'GEORGIA':
        source_hint = 'STATE'
    if source_hint not in SOURCE_OPTIONS:
        source_hint = source
    return {
        'terms': _dedupe_phrases(terms, limit=12),
        'query_variants': _dedupe_phrases(query_variants, limit=4),
        'related_policy_terms': _dedupe_phrases(related_policy_terms, limit=6),
        'source_hint': source_hint,
        'officer_brief': str(hit.get('officer_brief') or '').strip(),
    }


def _store_cached_ai_hint(query: str, source: str, hint: dict) -> None:
    key = f"{_normalize_query_key(query)}|{(source or 'ALL').upper()}"
    cache = _load_ai_hint_cache()
    cache[key] = {
        'terms': _dedupe_phrases(hint.get('terms') or [], limit=12),
        'query_variants': _dedupe_phrases(hint.get('query_variants') or [], limit=4),
        'related_policy_terms': _dedupe_phrases(hint.get('related_policy_terms') or [], limit=6),
        'source_hint': str(hint.get('source_hint') or source).strip().upper(),
        'officer_brief': str(hint.get('officer_brief') or '').strip(),
    }
    _write_ai_hint_cache(cache)


def _extract_json_object(text: str) -> dict:
    raw = (text or '').strip()
    if not raw:
        return {}
    fenced = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', raw, flags=re.S)
    if fenced:
        raw = fenced.group(1).strip()
    if not raw.startswith('{'):
        start = raw.find('{')
        end = raw.rfind('}')
        if start >= 0 and end > start:
            raw = raw[start:end + 1]
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, dict) else {}
    except json.JSONDecodeError:
        return {}


def _extract_json_payload(text: str):
    raw = (text or '').strip()
    if not raw:
        return None
    fenced = re.search(r'```(?:json)?\s*(\{.*?\}|\[.*?\])\s*```', raw, flags=re.S)
    if fenced:
        raw = fenced.group(1).strip()
    if not (raw.startswith('{') or raw.startswith('[')):
        start_obj = raw.find('{')
        start_arr = raw.find('[')
        start = min([i for i in (start_obj, start_arr) if i >= 0], default=-1)
        if start >= 0:
            tail = raw[start:]
            end_obj = tail.rfind('}')
            end_arr = tail.rfind(']')
            end = max(end_obj, end_arr)
            if end >= 0:
                raw = tail[:end + 1]
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        return None


def _ai_supplemental_candidates(query: str, source: str, results: list[LegalMatch]) -> list[dict]:
    if not (query or '').strip():
        return []
    existing_codes = {item.entry.code for item in results}
    prompt = (
        "You are assisting MCPD legal lookup when local corpus may be incomplete. "
        "Return STRICT JSON only in this format: "
        "{\"candidates\":[{\"source\":\"GEORGIA|UCMJ|BASE_ORDER|FEDERAL_USC\",\"code\":\"...\",\"title\":\"...\",\"reason\":\"...\",\"elements\":[\"...\"],"
        "\"minimum_punishment\":\"...\",\"maximum_punishment\":\"...\"}]}. "
        "Provide only likely, relevant candidates for the scenario. "
        "Do not include candidates already listed in this exclusion list.\n\n"
        f"Scenario: {query}\n"
        f"Current filter: {source}\n"
        f"Exclude codes already returned: {sorted(existing_codes)}"
    )
    answer = _safe_ai_guidance(prompt)
    if not answer:
        return []
    payload = _extract_json_payload(answer)
    if not isinstance(payload, dict):
        return []
    raw_candidates = payload.get('candidates')
    if not isinstance(raw_candidates, list):
        return []
    clean: list[dict] = []
    seen_codes: set[str] = set()
    for item in raw_candidates:
        if not isinstance(item, dict):
            continue
        code = str(item.get('code') or '').strip()
        title = str(item.get('title') or '').strip()
        src = str(item.get('source') or '').strip().upper()
        if src not in {'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
            continue
        if source == 'GEORGIA' and src != 'GEORGIA':
            continue
        if source == 'UCMJ' and src != 'UCMJ':
            continue
        if source == 'BASE_ORDER' and src != 'BASE_ORDER':
            continue
        if source == 'FEDERAL_USC' and src != 'FEDERAL_USC':
            continue
        if not code or not title:
            continue
        if code in existing_codes or code in seen_codes:
            continue
        seen_codes.add(code)
        elements = item.get('elements') if isinstance(item.get('elements'), list) else []
        clean.append({
            'source': src,
            'code': code,
            'title': title,
            'reason': str(item.get('reason') or '').strip(),
            'elements': [str(x).strip() for x in elements if str(x).strip()][:5],
            'minimum_punishment': str(item.get('minimum_punishment') or '').strip(),
            'maximum_punishment': str(item.get('maximum_punishment') or '').strip(),
        })
        if len(clean) >= 6:
            break
    return clean


def _plain_language_fact_profile(query: str) -> dict:
    normalized = _normalize_query_key(query)
    return {
        'public_sanitation': bool(re.search(r'public defecation|defecat\w+|poop\w+|feces|urinat\w+|peeing|public indecency|indecent exposure|lewd conduct', normalized)),
        'public_location': bool(re.search(r'\bpublic\b|street|sidewalk|road|parking lot|open area|outside', normalized)),
        'property_damage': bool(re.search(r'damag|destroy|destruction|vandal|graffiti|break|broke|smash', normalized)),
        'theft': bool(re.search(r'shoplift|steal|stole|stolen|theft|larceny|take|took|px theft|exchange theft', normalized)),
        'lawful_order': bool(re.search(r'lawful order|lawful command|direct order|disobey|refus\w+ command|article 92|regulation|ordered not to', normalized)),
        'entry_barment': bool(re.search(r'trespass|unauthorized entr|unlawful entr|barred|debarred|barment|returned to base|told not to return|ordered not to return|refus\w+ to leave|remain\w+ after', normalized)),
        'military_subject': bool(re.search(r'marine|soldier|airman|sailor|service member|military member|on duty|barracks|ucmj', normalized)),
    }


def _ai_candidate_relevance(query: str, jurisdiction: str, code: str, title: str, why_relevant: str, elements: list[str]) -> tuple[bool, str, str]:
    """Keep AI suggestions tied to the described conduct, not loose keywords.

    The AI sweep is useful for incomplete state/federal coverage, but officers
    should not see speculative federal/UCMJ candidates just because the query
    mentions a public place, base, street, or other background setting.
    """

    facts = _plain_language_fact_profile(query)
    candidate_text = _normalize_query_key(' '.join([code, title, why_relevant, ' '.join(elements)]))

    if facts['public_sanitation']:
        relevant_public_terms = re.search(
            r'disorderly|indecent|exposure|lewd|nuisance|sanitation|health|defecat|urination|public conduct',
            candidate_text,
        )
        if jurisdiction == 'STATE' and relevant_public_terms:
            return True, 'Strong Match', 'Matches the described public sanitation/indecency conduct and public-location facts.'
        if jurisdiction == 'UCMJ':
            military_public_terms = re.search(r'article 134|disorderly|indecent|prejudice|good order|conduct', candidate_text)
            if facts['military_subject'] and military_public_terms:
                return True, 'Related / Review If Needed', 'Possible military review only because the facts identify a military subject.'
            return False, '', 'Rejected UCMJ candidate because no military-status/order facts were described.'
        if jurisdiction == 'FEDERAL_USC':
            return False, '', 'Rejected federal candidate because public defecation does not describe federal property damage, theft, or barred-entry facts.'

    if re.search(r'1361|property damage|destruction', candidate_text) and not facts['property_damage']:
        return False, '', 'Rejected property-damage candidate because no damage/destruction facts were described.'
    if re.search(r'shoplift|exchange theft|px theft|theft|larceny', candidate_text) and not facts['theft']:
        return False, '', 'Rejected theft candidate because no taking/shoplifting facts were described.'
    if re.search(r'article 92|failure to obey|order or regulation|lawful order', candidate_text) and not facts['lawful_order']:
        return False, '', 'Rejected order/regulation candidate because no order/refusal facts were described.'
    if re.search(r'1382|reentry|military.*property|naval.*property|coast guard.*property|barred', candidate_text) and not facts['entry_barment']:
        return False, '', 'Rejected installation-entry candidate because no barred-entry/refusal-to-leave facts were described.'

    if jurisdiction == 'STATE':
        return True, 'Probable Match', 'State-law candidate may fit the described conduct; verify elements and current citation.'
    return True, 'Related / Review If Needed', 'Review only if the missing jurisdictional facts are confirmed.'


def _ai_multijurisdiction_candidates(query: str, source: str, state_code: str, results: list[LegalMatch]) -> list[dict]:
    if not (query or '').strip():
        return []
    state_name = STATE_LABELS.get(state_code, state_code)
    if source == 'BASE_ORDER':
        return []
    focus = {
        'ALL': f'{state_name} state law, federal United States Code, and UCMJ',
        'STATE': f'{state_name} state law',
        'UCMJ': 'UCMJ punitive articles',
        'FEDERAL_USC': 'federal United States Code',
    }.get(source, f'{state_name} state law, federal United States Code, and UCMJ')
    existing_codes = {item.entry.code for item in results}
    prompt = (
        "You are assisting MCPD legal lookup. Return STRICT JSON only. "
        "Do not provide legal advice or final charging decisions. "
        "Find likely relevant law candidates for officer review and official verification. "
        "Read the full scenario before choosing candidates. Identify the conduct, location, likely legal category, "
        "and jurisdictional facts. Suppress candidates based only on loose keywords. "
        "Do not include property-damage, theft, federal-entry, or Article 92/order candidates unless the scenario "
        "actually describes damage, taking, barred/refused entry, or order/regulation facts. "
        "Use this schema exactly: "
        "{\"candidates\":[{\"jurisdiction\":\"STATE|FEDERAL_USC|UCMJ\",\"state\":\"AA\",\"code\":\"...\","
        "\"title\":\"...\",\"why_relevant\":\"...\",\"elements\":[\"...\"],\"verification_note\":\"...\"}]}. "
        "Only include candidates that are plausibly relevant to the scenario and focus. "
        "If unsure about an exact citation, say so in verification_note instead of inventing certainty.\n\n"
        f"Scenario: {query}\n"
        f"Selected state: {state_name} ({state_code})\n"
        f"Search focus: {focus}\n"
        f"Exclude local corpus codes already returned: {sorted(existing_codes)}"
    )
    answer = _safe_ai_guidance(prompt)
    if not answer:
        return []
    payload = _extract_json_payload(answer)
    if not isinstance(payload, dict):
        return []
    raw_candidates = payload.get('candidates')
    if not isinstance(raw_candidates, list):
        return []
    allowed = {'STATE', 'FEDERAL_USC', 'UCMJ'}
    if source == 'STATE':
        allowed = {'STATE'}
    elif source == 'FEDERAL_USC':
        allowed = {'FEDERAL_USC'}
    elif source == 'UCMJ':
        allowed = {'UCMJ'}
    clean = []
    seen = set()
    for item in raw_candidates:
        if not isinstance(item, dict):
            continue
        jurisdiction = str(item.get('jurisdiction') or item.get('source') or '').strip().upper()
        if jurisdiction in {'FEDERAL', 'USC'}:
            jurisdiction = 'FEDERAL_USC'
        if jurisdiction not in allowed:
            continue
        code = str(item.get('code') or '').strip()
        title = str(item.get('title') or '').strip()
        if not code or not title:
            continue
        key = (jurisdiction, code.lower())
        if key in seen or code in existing_codes:
            continue
        why_relevant = str(item.get('why_relevant') or item.get('reason') or '').strip()
        elements = _dedupe_phrases(item.get('elements') if isinstance(item.get('elements'), list) else [], limit=6)
        allowed_candidate, tier, gate_note = _ai_candidate_relevance(query, jurisdiction, code, title, why_relevant, list(elements))
        if not allowed_candidate:
            continue
        seen.add(key)
        clean.append(
            {
                'jurisdiction': jurisdiction,
                'state': state_code if jurisdiction == 'STATE' else '',
                'label': f'{state_name} State Law' if jurisdiction == 'STATE' else _source_label(jurisdiction),
                'code': code,
                'title': title,
                'why_relevant': why_relevant,
                'matched_facts': gate_note,
                'tier': tier,
                'elements': elements,
                'verification_note': str(item.get('verification_note') or 'Verify against the current official source before use.').strip(),
            }
        )
        if len(clean) >= 8:
            break
    return clean


def _ai_search_hints(query: str, source: str, results: list[LegalMatch] | None = None) -> dict:
    clean_query = _normalize_query_key(query)
    if not clean_query:
        return {'terms': [], 'query_variants': [], 'related_policy_terms': [], 'source_hint': source, 'officer_brief': ''}
    cached = _get_cached_ai_hint(clean_query, source)
    if cached:
        return cached
    candidate_summary = _result_summary(results or [], limit=6) if results else 'No strong local legal candidates yet.'
    prompt = (
        "You are improving retrieval for MCPD legal lookup. "
        "You are NOT allowed to invent citations or claim a statute exists if it is not already in the local corpus. "
        "Your task is only to produce grounded search hints and a short officer-facing brief. "
        "Return STRICT JSON with keys: terms (array), query_variants (array), related_policy_terms (array), "
        "source_hint (ALL|STATE|UCMJ|BASE_ORDER|FEDERAL_USC), officer_brief (string). "
        "Keep every phrase short and operational. Use related_policy_terms only for approved local orders/policy cross-reference searching. "
        "Do not include any text outside JSON.\n\n"
        f"Scenario: {clean_query}\n"
        f"Current filter: {source}\n"
        f"Current local candidates:\n{candidate_summary}"
    )
    answer = _safe_ai_guidance(prompt)
    if not answer:
        return {'terms': [], 'query_variants': [], 'related_policy_terms': [], 'source_hint': source, 'officer_brief': ''}
    parsed = _extract_json_object(answer)
    terms = _dedupe_phrases(parsed.get('terms') if isinstance(parsed.get('terms'), list) else [], limit=12)
    query_variants = _dedupe_phrases(parsed.get('query_variants') if isinstance(parsed.get('query_variants'), list) else [], limit=4)
    related_policy_terms = _dedupe_phrases(parsed.get('related_policy_terms') if isinstance(parsed.get('related_policy_terms'), list) else [], limit=6)
    source_hint = str(parsed.get('source_hint') or source).strip().upper()
    if source_hint == 'GEORGIA':
        source_hint = 'STATE'
    if source_hint not in SOURCE_OPTIONS:
        source_hint = source
    result = {
        'terms': terms,
        'query_variants': query_variants,
        'related_policy_terms': related_policy_terms,
        'source_hint': source_hint,
        'officer_brief': str(parsed.get('officer_brief') or '').strip(),
    }
    if any(result.get(key) for key in ('terms', 'query_variants', 'related_policy_terms', 'officer_brief')):
        _store_cached_ai_hint(clean_query, source, result)
    return result


def _merge_ai_results(base_results: list[LegalMatch], ai_results: list[LegalMatch]) -> list[LegalMatch]:
    merged: dict[str, LegalMatch] = {item.entry.code: item for item in base_results}
    top_base_score = base_results[0].score if base_results else 0
    min_ai_new_score = int(top_base_score * 0.92) if top_base_score else 0
    for item in ai_results:
        existing = merged.get(item.entry.code)
        if existing is None:
            if top_base_score and item.score < min_ai_new_score:
                continue
            merged[item.entry.code] = LegalMatch(
                entry=item.entry,
                score=item.score + 6,
                reasons=tuple(dict.fromkeys(tuple(item.reasons) + ('AI scenario expansion match',))),
            )
            continue
        ai_score = item.score + 6
        if ai_score > existing.score:
            merged[item.entry.code] = LegalMatch(
                entry=item.entry,
                score=ai_score,
                reasons=tuple(dict.fromkeys(tuple(existing.reasons) + tuple(item.reasons) + ('AI scenario expansion match',))),
            )
        else:
            merged[item.entry.code] = LegalMatch(
                entry=existing.entry,
                score=existing.score,
                reasons=tuple(dict.fromkeys(tuple(existing.reasons) + ('AI scenario expansion considered',))),
            )
    ordered = sorted(merged.values(), key=lambda item: (-item.score, item.entry.source, item.entry.code))
    return ordered[:25]


_NARRATIVE_CONNECTORS = frozenset({
    'was', 'were', 'is', 'are', 'be', 'been', 'the', 'a', 'an',
    'he', 'she', 'they', 'we', 'i', 'it', 'his', 'her', 'their',
    'and', 'then', 'while', 'when', 'that', 'which', 'who', 'whom',
    'in', 'on', 'at', 'to', 'of', 'for', 'with', 'by', 'from',
    'after', 'before', 'but', 'because', 'not', 'did', 'had', 'has',
    'my', 'me', 'him', 'us', 'them', 'into', 'over', 'out', 'up',
    'this', 'that', 'these', 'those', 'so', 'if', 'also', 'no',
})

# Words that describe the physical setting of an incident but are NOT offense
# elements themselves.  When these appear in a multi-word query the keyword
# engine pulls in traffic/property laws purely because of the location word,
# not because the described incident is traffic-related.
# Conservative list: only words that almost never double as offense elements.
# (road, street, lane, parking, yard, vehicle, school, store are deliberately
# excluded because they also appear in compound offense names.)
_LOCATION_ONLY_WORDS = frozenset({
    'roadway', 'freeway', 'expressway', 'interstate',
    'intersection', 'crosswalk', 'sidewalk', 'median',
    'overpass', 'underpass', 'curb', 'curbside', 'pavement',
    'hallway', 'corridor', 'stairwell', 'lobby',
    'vicinity', 'scene',
})


def _is_narrative_query(query: str) -> bool:
    words = (query or '').strip().split()
    if len(words) < 5:
        return False
    lower_words = {re.sub(r"[^a-z']", '', w.lower()) for w in words}
    return len(lower_words & _NARRATIVE_CONNECTORS) >= 2


def _strip_location_context_words(query: str) -> str:
    """Remove pure-setting words so they don't drive keyword offense matching."""
    words = (query or '').split()
    filtered = [w for w in words if re.sub(r"[^a-z]", '', w.lower()) not in _LOCATION_ONLY_WORDS]
    return ' '.join(filtered)


def _ai_interpret_narrative_query(query: str, source: str) -> dict:
    clean_key = _normalize_query_key(query)
    if not clean_key:
        return {}
    cache_key = f"narr|{clean_key}|{(source or 'ALL').upper()}"
    try:
        if NARRATIVE_INTERP_CACHE_PATH.exists():
            cached_data = json.loads(NARRATIVE_INTERP_CACHE_PATH.read_text(encoding='utf-8'))
            if isinstance(cached_data, dict) and cache_key in cached_data:
                return cached_data[cache_key]
    except Exception:
        pass
    prompt = (
        "You are a law enforcement legal reference assistant. "
        "An officer described an incident in plain language. "
        "Read the ENTIRE statement and identify the legal concepts it describes. "
        "Do NOT invent statute numbers or cite specific codes. "
        "Return STRICT JSON only, no other text:\n"
        "{\n"
        '  "primary_search": "<3-7 word legal search phrase capturing the core offense>",\n'
        '  "additional_terms": ["<secondary offense or element>", "<another term if applicable>"],\n'
        '  "interpretation_note": "<one sentence describing the legal scenario you understood>"\n'
        "}\n\n"
        f"Incident description: {clean_key}\n"
        f"Jurisdiction focus: {source}"
    )
    answer = _safe_ai_guidance(prompt)
    if not answer:
        return {}
    parsed = _extract_json_object(answer)
    if not isinstance(parsed, dict):
        return {}
    primary_search = str(parsed.get('primary_search') or '').strip()
    if not primary_search or len(primary_search.split()) < 2:
        return {}
    result = {
        'primary_search': primary_search,
        'additional_terms': _dedupe_phrases(
            parsed.get('additional_terms') if isinstance(parsed.get('additional_terms'), list) else [],
            limit=4,
        ),
        'interpretation_note': str(parsed.get('interpretation_note') or '').strip(),
    }
    try:
        NARRATIVE_INTERP_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
        existing_cache = {}
        if NARRATIVE_INTERP_CACHE_PATH.exists():
            try:
                existing_cache = json.loads(NARRATIVE_INTERP_CACHE_PATH.read_text(encoding='utf-8'))
            except Exception:
                pass
        if not isinstance(existing_cache, dict):
            existing_cache = {}
        existing_cache[cache_key] = result
        NARRATIVE_INTERP_CACHE_PATH.write_text(json.dumps(existing_cache, indent=2), encoding='utf-8')
    except Exception:
        pass
    return result


def _normalize_source_label(value: str, fallback: str = 'ALL') -> str:
    raw = (value or '').strip().upper()
    if raw in {'GA', 'GEORGIA', 'GEORGIA_CODE'}:
        return 'GEORGIA'
    if raw in {'UCMJ', 'MCM', 'ARTICLE'}:
        return 'UCMJ'
    if raw in {'BASE_ORDER', 'BASE', 'ORDER', 'MCLBAO'}:
        return 'BASE_ORDER'
    if raw in {'FEDERAL_USC', 'FEDERAL', 'USC', 'UNITED_STATES_CODE'}:
        return 'FEDERAL_USC'
    if raw in {'ALL', ''}:
        return fallback
    return fallback


def _split_multi(value: str) -> list[str]:
    text = (value or '').strip()
    if not text:
        return []
    for sep in ('|', ';', ','):
        if sep in text:
            return [item.strip() for item in text.split(sep) if item.strip()]
    return [text]


def _rows_to_payload(rows: list[dict], target_source: str) -> dict:
    georgia_entries = []
    ucmj_entries = []
    base_order_entries = []
    federal_usc_entries = []
    for row in rows:
        code = (row.get('code') or '').strip()
        title = (row.get('title') or '').strip()
        summary = (row.get('summary') or '').strip()
        if not (code and title and summary):
            continue
        row_source = _normalize_source_label(row.get('source', ''), target_source if target_source != 'ALL' else 'GEORGIA')
        if target_source in {'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
            row_source = target_source
        entry = {
            'source': row_source,
            'code': code,
            'title': title,
            'summary': summary,
            'elements': _split_multi(row.get('elements', '')),
            'notes': (row.get('notes') or '').strip(),
            'keywords': _split_multi(row.get('keywords', '')),
            'related_codes': _split_multi(row.get('related_codes', '')),
            'minimum_punishment': (row.get('minimum_punishment') or '').strip(),
            'maximum_punishment': (row.get('maximum_punishment') or '').strip(),
            'aliases': _split_multi(row.get('aliases', '')),
            'synonyms': _split_multi(row.get('synonyms', '')),
            'narrative_triggers': _split_multi(row.get('narrative_triggers', '')),
            'lesser_included_offenses': _split_multi(row.get('lesser_included_offenses', '')),
            'alternative_offenses': _split_multi(row.get('alternative_offenses', '')),
            'category': (row.get('category') or '').strip(),
            'subcategory': (row.get('subcategory') or '').strip(),
            'severity': (row.get('severity') or '').strip(),
            'source_reference': (row.get('source_reference') or '').strip(),
        }
        if row_source == 'UCMJ':
            ucmj_entries.append(entry)
        elif row_source == 'BASE_ORDER':
            base_order_entries.append(entry)
        elif row_source == 'FEDERAL_USC':
            federal_usc_entries.append(entry)
        else:
            georgia_entries.append(entry)
    return {
        'georgia_codes': georgia_entries,
        'ucmj_articles': ucmj_entries,
        'base_orders': base_order_entries,
        'federal_usc_codes': federal_usc_entries,
    }


def _read_tabular_upload(upload) -> list[dict]:
    filename = (upload.filename or '').strip()
    ext = Path(filename).suffix.lower()
    if ext == '.csv':
        text = upload.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        return [{(k or '').strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]
    if ext in {'.xlsx', '.xlsm'}:
        if load_workbook is None:
            raise RuntimeError('openpyxl is required for Excel imports')
        upload.seek(0)
        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [(str(cell or '').strip().lower()) for cell in rows[0]]
        out: list[dict] = []
        for row in rows[1:]:
            record = {}
            for i, key in enumerate(headers):
                if not key:
                    continue
                value = row[i] if i < len(row) else ''
                record[key] = '' if value is None else str(value).strip()
            if any(record.values()):
                out.append(record)
        return out
    raise ValueError('Unsupported file type')


def _result_summary(results, limit=5):
    lines = []
    for item in results[:limit]:
        lines.append(
            f"- {item.entry.code} | {item.entry.title} | source={item.entry.source} | elements={' ; '.join(item.entry.elements[:3])} | minimum={item.entry.minimum_punishment or 'N/A'} | maximum={item.entry.maximum_punishment or 'N/A'}"
        )
    return '\n'.join(lines)


def _order_reference_matches(query: str, related_terms: tuple[str, ...] | list[str] = (), limit: int = 8) -> list[dict]:
    clean_query = (query or '').strip()
    if not clean_query:
        return []
    from .orders import search_orders_with_ai_assist

    combined_query = ' '.join(
        part for part in [clean_query, ' '.join(_dedupe_phrases(related_terms, limit=4))] if part.strip()
    ).strip()
    documents, _strategy = search_orders_with_ai_assist(combined_query, status_filter='ACTIVE')
    matches = []
    query_terms = {
        term for term in _tokenize(clean_query)
        if len(term) >= 3 and term not in STOPWORDS
    }
    for document in documents:
        reasons = list(getattr(document, 'match_reasons', []) or [])[:4]
        searchable_text = ' '.join((
            document.title or '',
            document.summary or '',
            document.match_snippet or '',
            document.order_number or '',
            document.memo_number or '',
        ))
        document_terms = set(_tokenize(searchable_text))
        meaningful_overlap = query_terms & document_terms
        stopword_only_reason = bool(reasons) and not meaningful_overlap and all(
            re.fullmatch(r'matched "[a-z]+" in the title', reason.strip().lower())
            for reason in reasons
        )
        if stopword_only_reason:
            continue
        if not meaningful_overlap and getattr(document, 'search_confidence', 0) < 70:
            continue
        matches.append(
            {
                'id': document.id,
                'title': document.title,
                'source_type': document.source_type or 'LOCAL_DOCUMENT',
                'order_number': document.order_number or document.memo_number or '',
                'snippet': document.match_snippet or document.summary or '',
                'confidence': getattr(document, 'search_confidence', 0),
                'reasons': reasons,
                'download_available': bool(getattr(document, 'download_available', False)),
            }
        )
        if len(matches) >= limit:
            break
    return matches


def _group_results(results: list[LegalMatch]) -> list[dict]:
    buckets = {'GEORGIA': [], 'UCMJ': [], 'BASE_ORDER': [], 'FEDERAL_USC': []}
    for item in results:
        source = item.entry.source if item.entry.source in buckets else 'GEORGIA'
        buckets[source].append(item)

    grouped = []
    for source in ('GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'):
        rows = buckets[source]
        if not rows:
            continue
        grouped.append(
            {
                'source': source,
                'label': {
                    'GEORGIA': 'Georgia Code',
                    'UCMJ': 'UCMJ',
                    'BASE_ORDER': 'MCLB Albany Base Orders',
                    'FEDERAL_USC': 'United States Code',
                }[source],
                'strongest': [item for item in rows if item.certainty_bucket == 'strong'],
                'probable': [item for item in rows if item.certainty_bucket == 'probable'],
                'possible': [item for item in rows if item.certainty_bucket == 'possible'],
            }
        )
    return grouped


def _source_label(value: str) -> str:
    return {
        'ALL': 'All Sources',
        'GEORGIA': 'Georgia Code',
        'UCMJ': 'UCMJ',
        'BASE_ORDER': 'MCLB Albany Base Orders',
        'FEDERAL_USC': 'United States Code',
    }.get((value or '').upper(), value or 'Unknown')


def _is_code_lookup_query(query: str) -> bool:
    q = (query or '').lower()
    return bool(
        re.search(r'\b(?:ocga\s*)?\d{1,2}-\d{1,2}(?:-\d{1,3}(?:\.\d+)?)?\b', q)
        or re.search(r'\barticle\s+\d{1,3}[a-z]?\b', q)
        or re.search(r'\b\d+\s*usc\s*§?\s*\d+', q)
    )


def _filter_results_for_display(query: str, source: str, results: list[LegalMatch], include_possible: bool = False) -> list[LegalMatch]:
    if not query or not results:
        return results
    if include_possible:
        return results

    code_lookup_like = _is_code_lookup_query(query)
    signal_reasons = {
        'Exact code match',
        'Code number match',
        'Article number match',
        'Exact title match',
        'Exact keyword match',
        'Keyword phrase match',
        'Intent phrase match',
        'Intent alignment match',
    }
    confident = [
        item for item in results
        if item.certainty_bucket in {'strong', 'probable'}
        and (
            code_lookup_like
            or len(item.matched_terms) >= 1
            or item.confidence >= 82
            or any(reason in signal_reasons for reason in item.reasons)
        )
    ]
    if confident:
        return confident[:30] if source in {'GEORGIA', 'STATE'} else confident[:14]

    floor = 46 if code_lookup_like else 50
    narrowed = [
        item for item in results
        if item.confidence >= floor
        and (
            code_lookup_like
            or len(item.matched_terms) >= 1
            or item.confidence >= 78
            or any(reason in signal_reasons for reason in item.reasons)
        )
    ]
    if narrowed:
        return narrowed[:8]
    return results[:5]


def _render_legal_lookup(default_source='ALL'):
    query = (request.args.get('q') or '').strip()
    user_preferred_state = getattr(current_user, 'preferred_legal_state', None) or ''
    state = _normalize_state(
        request.args.get('state')
        or session.get('legal_last_state')
        or user_preferred_state
        or 'GA'
    )
    source = _normalize_lookup_source(request.args.get('source') or default_source or 'ALL')
    include_possible = (request.args.get('show_possible') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    normalized_query = query.lower()
    speed_phrase = bool(re.search(r'\b\d{1,3}\s*(?:in|/)\s*(?:a\s*)?\d{1,3}\b', normalized_query))
    deterministic_lookup = speed_phrase
    ai_expansion_enabled = bool(current_app.config.get('LEGAL_AI_EXPANSION_ENABLED', False))
    code_lookup_like = bool(re.search(r'\b(?:ocga\s*)?\d{1,2}-\d{1,2}(?:-\d{1,3}(?:\.\d+)?)?\b', query.lower()))
    ai_hints = {'terms': [], 'query_variants': [], 'related_policy_terms': [], 'source_hint': source, 'officer_brief': ''}
    ai_interpretation_note = ''

    # --- Baseline keyword search ---
    # For queries of ≥3 words that contain pure location/setting words (roadway,
    # sidewalk, etc.), strip those words before the keyword pass so the engine
    # matches on the actual offense terms, not the physical location described.
    search_query = query
    if (
        query
        and not deterministic_lookup
        and not code_lookup_like
        and len(query.split()) >= 3
    ):
        stripped = _strip_location_context_words(query)
        if stripped and len(stripped.split()) >= 2 and stripped.lower() != query.lower():
            search_query = stripped

    raw_results = _search_entries_for_scope(search_query, source, state)

    # --- AI narrative interpretation ---
    # For full natural-language sentences, read the ENTIRE statement semantically
    # and use AI-interpreted legal terms as the PRIMARY result set.  The AI results
    # REPLACE (not augment) the keyword results for narrative queries so that
    # location words in the original query never re-introduce mismatched law hits.
    allow_narrative_interp = bool(
        ai_expansion_enabled
        and query
        and not deterministic_lookup
        and not code_lookup_like
        and _is_narrative_query(query)
    )
    if allow_narrative_interp:
        interp = _ai_interpret_narrative_query(query, source)
        primary_search = (interp.get('primary_search') or '').strip()
        if primary_search and primary_search.lower() not in (query.lower(), search_query.lower()):
            ai_interpretation_note = interp.get('interpretation_note', '')
            interp_results = _search_entries_for_scope(primary_search, source, state)
            for extra_term in (interp.get('additional_terms') or [])[:3]:
                extra_term = (extra_term or '').strip()
                if extra_term and extra_term.lower() != primary_search.lower():
                    extra_results = _search_entries_for_scope(extra_term, source, state)
                    interp_results = _merge_ai_results(interp_results, extra_results)
            # AI results are the authority for narrative queries; keyword results
            # supplement only if they already appear in the AI result set
            ai_codes = {item.entry.code for item in interp_results}
            keyword_supplement = [r for r in raw_results if r.entry.code in ai_codes]
            raw_results = _merge_ai_results(interp_results, keyword_supplement)

    results = _filter_results_for_display(query, source, raw_results, include_possible=include_possible)

    allow_ai_expansion = bool(
        ai_expansion_enabled
        and query
        and not deterministic_lookup
        and not code_lookup_like
        and len(query.split()) >= 2
    )
    if allow_ai_expansion:
        ai_hints = _ai_search_hints(query, source, raw_results)
        boosted_source = source if source != 'ALL' else ai_hints['source_hint']
        variant_queries = _dedupe_phrases(
            list(ai_hints.get('query_variants') or [])
            + ([f"{query} {' '.join((ai_hints.get('terms') or [])[:4])}".strip()] if ai_hints.get('terms') else []),
            limit=4,
        )
        for expanded_query in variant_queries:
            if expanded_query.lower() == query.lower():
                continue
            boosted_results = _search_entries_for_scope(expanded_query, boosted_source, state)
            raw_results = _merge_ai_results(raw_results, boosted_results)
        results = _filter_results_for_display(query, source, raw_results, include_possible=include_possible)
    page_title = {
        'STATE': f'{STATE_LABELS.get(state, state)} Law Lookup',
        'UCMJ': 'UCMJ Lookup',
        'BASE_ORDER': 'Base Order Lookup',
        'FEDERAL_USC': 'United States Code Lookup',
        'ALL': 'Law Lookup',
    }[source]
    page_subtitle = {
        'STATE': f'Search {STATE_LABELS.get(state, state)} state-law candidates and verified local state corpus when available.',
        'UCMJ': 'Search punitive articles with structured offense elements.',
        'BASE_ORDER': 'Search MCLB Albany base-order references and punitive language.',
        'FEDERAL_USC': 'Search United States Code offense references used in law-enforcement contexts.',
        'ALL': f'Search {STATE_LABELS.get(state, state)} state law, United States Code, UCMJ, and MCLB Albany base orders in one place.',
    }[source]
    example_terms = {
        'STATE': (
            ('65 in a 25', 'STATE'),
            ('reckless driving', 'STATE'),
            ('following too close', 'STATE'),
            ('dui refusal', 'STATE'),
            ('domestic battery', 'STATE'),
        ),
        'UCMJ': (
            ('AWOL', 'UCMJ'),
            ('disrespect', 'UCMJ'),
            ('failure to obey order', 'UCMJ'),
            ('assault', 'UCMJ'),
        ),
        'BASE_ORDER': (
            ('underage drinking on base', 'BASE_ORDER'),
            ('drove on base while intoxicated', 'BASE_ORDER'),
            ('barracks disturbance', 'BASE_ORDER'),
            ('base traffic violation', 'BASE_ORDER'),
        ),
        'FEDERAL_USC': (
            ('felon in possession federal', 'FEDERAL_USC'),
            ('stole government property', 'FEDERAL_USC'),
            ('damaged government property', 'FEDERAL_USC'),
            ('18 usc 922g', 'FEDERAL_USC'),
        ),
        'ALL': (
            ('65 in a 25', 'STATE'),
            ('reckless driving', 'STATE'),
            ('AWOL', 'UCMJ'),
            ('disrespect', 'UCMJ'),
        ),
    }[source]
    previous_query = str(session.get('legal_last_query') or '').strip()
    if current_app.config.get('LEGAL_QUERY_LOG_ENABLED', True):
        _append_legal_query_log(query, source, results)
        _append_search_failure_log(query, source, results, previous_query=previous_query)
    if query:
        session['legal_last_query'] = query
        session['legal_last_state'] = state
        if getattr(current_user, 'is_authenticated', False) and getattr(current_user, 'preferred_legal_state', None) != state:
            try:
                current_user.preferred_legal_state = state
                db.session.commit()
            except Exception:
                db.session.rollback()
    lead_result = results[0] if results else None
    grouped_results = _group_results(results[1:] if len(results) > 1 else [])
    order_reference_matches = _order_reference_matches(query, ai_hints.get('related_policy_terms', ()))
    needs_ai_candidates = bool(
        query
        and (
            not results
            or not (state == 'GA')
            or not lead_result
            or lead_result.confidence < 75
        )
    )
    ai_candidates = _ai_multijurisdiction_candidates(query, source, state, results) if needs_ai_candidates else []
    overlap_note = ''
    ai_brief = (ai_hints.get('officer_brief') or '').strip()
    active_sources = [group['source'] for group in grouped_results if (group.get('strongest') or group.get('probable') or group.get('possible'))]
    if len(active_sources) >= 2:
        overlap_note = 'Multiple jurisdictional sources may apply - verify prosecutive authority, command authority, and legal review.'
    return _no_store_response(
        render_template(
            'legal_lookup.html',
            title=page_title,
            user=current_user,
            query=query,
            source=source,
            state=state,
            state_options=STATE_OPTIONS,
            state_label=STATE_LABELS.get(state, state),
            results=results,
            grouped_results=grouped_results,
            source_options=SOURCE_OPTIONS,
            page_title=page_title,
            page_subtitle=page_subtitle,
            example_terms=example_terms,
            search_tip='Search by code number, article number, title, offense phrase, or plain-language incident narrative.',
            corpus_status=corpus_status(),
            ai_terms=ai_hints.get('terms', ()),
            ai_query_variants=ai_hints.get('query_variants', ()),
            ai_related_policy_terms=ai_hints.get('related_policy_terms', ()),
            ai_candidates=ai_candidates,
            ai_brief=ai_brief,
            ai_interpretation_note=ai_interpretation_note,
            state_corpus_available=(state == 'GA'),
            code_lookup_like=code_lookup_like,
            show_possible=include_possible,
            overlap_note=overlap_note,
            order_reference_matches=order_reference_matches,
            lead_result=lead_result,
            lead_download_info=reference_download_info(lead_result.entry) if lead_result else {'available': False},
            reference_download_info=reference_download_info,
        )
    )


def _current_search(default_source='ALL'):
    query = (request.args.get('q') or '').strip()
    user_preferred_state = getattr(current_user, 'preferred_legal_state', None) or ''
    state = _normalize_state(
        request.args.get('state')
        or session.get('legal_last_state')
        or user_preferred_state
        or 'GA'
    )
    source = _normalize_lookup_source(request.args.get('source') or default_source or 'ALL')
    return query, source, _search_entries_for_scope(query, source, state)


@bp.route('/legal')
@login_required
def legal_home():
    return redirect(url_for('legal.legal_lookup'))


@bp.route('/legal/search')
@login_required
def legal_lookup():
    return _render_legal_lookup('ALL')


@bp.route('/legal/georgia')
@login_required
def georgia_code_lookup():
    return redirect(url_for('legal.legal_lookup', source='STATE', state='GA'))


@bp.route('/legal/ucmj')
@login_required
def ucmj_lookup():
    return redirect(url_for('legal.legal_lookup', source='UCMJ'))


@bp.route('/legal/base-orders')
@login_required
def base_order_lookup():
    return redirect(url_for('legal.legal_lookup', source='BASE_ORDER'))


@bp.route('/legal/federal-usc')
@login_required
def federal_usc_lookup():
    return redirect(url_for('legal.legal_lookup', source='FEDERAL_USC'))


@bp.route('/legal/reference/<source>/<path:code>')
@login_required
def view_reference(source, code):
    source = _legal_source_or_404(source)
    entry = get_entry(source, code)
    if not entry:
        abort(404)
    query = (request.args.get('q') or '').strip()
    match_context = _lookup_match_context(query, source, entry.code)
    download_info = reference_download_info(entry)
    return _no_store_response(
        render_template(
            'legal_reference_view.html',
            title=f'{entry.code} | {entry.title}',
            user=current_user,
            entry=entry,
            query=query,
            match_context=match_context,
            download_info=download_info,
        )
    )


@bp.route('/legal/reference/<source>/<path:code>/download')
@login_required
def download_reference(source, code):
    source = _legal_source_or_404(source)
    entry = get_entry(source, code)
    if not entry:
        abort(404)
    download_info = reference_download_info(entry)
    if not download_info.get('available'):
        abort(404)
    if download_info.get('mode') == 'file':
        file_name = download_info.get('file_name') or f'{entry.code}.txt'
        files = _source_file_candidates(entry)
        if files:
            return send_file(files[0], as_attachment=True, download_name=file_name)
        abort(404)
    content = _reference_text(entry).encode('utf-8')
    return send_file(
        io.BytesIO(content),
        as_attachment=True,
        download_name=download_info.get('file_name') or f'{entry.code}.txt',
        mimetype=download_info.get('mime') or 'text/plain',
    )


@bp.route('/legal/debug')
@login_required
def legal_debug():
    if not current_user.can_manage_site():
        flash('Only the Website Controller can use search debug.', 'danger')
        return redirect(url_for('legal.legal_lookup'))

    query = (request.args.get('q') or '').strip()
    source = (request.args.get('source') or 'ALL').strip().upper()
    if source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
        source = 'ALL'

    include_possible = (request.args.get('show_possible') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    strict_results = search_entries(query, source, strict_gating=True) if query else []
    relaxed_results = search_entries(query, source, strict_gating=False) if query else []
    displayed_results = _filter_results_for_display(query, source, strict_results, include_possible=include_possible) if query else []

    displayed_codes = {item.entry.code for item in displayed_results}
    strict_codes = {item.entry.code for item in strict_results}
    relaxed_codes = {item.entry.code for item in relaxed_results}

    hidden_from_display = [item for item in strict_results if item.entry.code not in displayed_codes]
    only_in_relaxed = [item for item in relaxed_results if item.entry.code not in strict_codes]

    return _no_store_response(
        render_template(
            'admin_legal_debug.html',
            title='Legal Search Debug',
            user=current_user,
            query=query,
            source=source,
            source_options=('ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'),
            source_label=_source_label(source),
            show_possible=include_possible,
            strict_results=strict_results[:120],
            displayed_results=displayed_results[:120],
            hidden_from_display=hidden_from_display[:120],
            only_in_relaxed=only_in_relaxed[:120],
            tuning_cases=_load_legal_tuning_cases(limit=60),
        )
    )


@bp.route('/legal/debug/save-case', methods=['POST'])
@login_required
def legal_debug_save_case():
    if not current_user.can_manage_site():
        flash('Only the Website Controller can save tuning cases.', 'danger')
        return redirect(url_for('legal.legal_lookup'))

    query = (request.form.get('q') or '').strip()
    source = (request.form.get('source') or 'ALL').strip().upper()
    if source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
        source = 'ALL'

    raw_expected = (request.form.get('expected_codes') or '').strip()
    notes = (request.form.get('notes') or '').strip()
    expected_codes = [part.strip() for part in re.split(r'[,;\n|]+', raw_expected) if part.strip()]

    if not query:
        flash('Enter a query before saving a tuning case.', 'warning')
        return redirect(url_for('legal.legal_debug', q=query, source=source))
    if not expected_codes:
        flash('Enter at least one expected code (comma separated).', 'warning')
        return redirect(url_for('legal.legal_debug', q=query, source=source))

    _append_legal_tuning_case(query, source, expected_codes, notes)
    flash('Tuning case saved.', 'success')
    return redirect(url_for('legal.legal_debug', q=query, source=source, show_possible='1'))


@bp.route('/legal/debug/delete-case', methods=['POST'])
@login_required
def legal_debug_delete_case():
    if not current_user.can_manage_site():
        flash('Only the Website Controller can delete tuning cases.', 'danger')
        return redirect(url_for('legal.legal_lookup'))

    case_id = (request.form.get('case_id') or '').strip()
    query = (request.form.get('q') or '').strip()
    source = (request.form.get('source') or 'ALL').strip().upper()
    if source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
        source = 'ALL'

    if not case_id:
        flash('Missing case id.', 'warning')
    elif _delete_legal_tuning_case(case_id):
        flash('Tuning case deleted.', 'success')
    else:
        flash('Tuning case not found or could not be deleted.', 'warning')
    return redirect(url_for('legal.legal_debug', q=query, source=source, show_possible='1'))


@bp.route('/legal/export.json')
@login_required
def legal_lookup_export_json():
    query, source, results = _current_search('ALL')
    payload = {
        'query': query,
        'source': source,
        'count': len(results),
        'results': [
            {
                'source': item.entry.source,
                'code': item.entry.code,
                'title': item.entry.title,
                'summary': item.entry.summary,
                'elements': list(item.entry.elements),
                'notes': item.entry.notes,
                'keywords': list(item.entry.keywords),
                'related_codes': list(item.entry.related_codes),
                'minimum_punishment': item.entry.minimum_punishment,
                'maximum_punishment': item.entry.maximum_punishment,
                'score': item.score,
                'confidence': item.confidence,
                'matched_terms': list(item.matched_terms),
                'reasons': list(item.reasons),
                'warning': item.warning,
                'certainty_bucket': item.certainty_bucket,
            }
            for item in results
        ],
    }
    return Response(
        response=json.dumps(payload),
        status=200,
        mimetype='application/json',
        headers={
            'Content-Disposition': 'attachment; filename=legal-lookup.json',
            'Cache-Control': 'no-store',
        },
    )


@bp.route('/legal/export.csv')
@login_required
def legal_lookup_export_csv():
    query, source, results = _current_search('ALL')
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['source', 'code', 'title', 'summary', 'elements', 'minimum_punishment', 'maximum_punishment', 'notes', 'keywords', 'query', 'filter'])
    for item in results:
        writer.writerow([
            item.entry.source,
            item.entry.code,
            item.entry.title,
            item.entry.summary,
            ' | '.join(item.entry.elements),
            item.entry.minimum_punishment,
            item.entry.maximum_punishment,
            item.entry.notes,
            ' | '.join(item.entry.keywords),
            query,
            source,
        ])
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=legal-lookup.csv',
            'Cache-Control': 'no-store',
        },
    )


@bp.route('/legal/corpus/export.json')
@login_required
def legal_corpus_export_json():
    source = (request.args.get('source') or 'ALL').strip().upper()
    if source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
        source = 'ALL'
    payload = export_corpus_payload(source)
    return Response(
        response=json.dumps(payload),
        status=200,
        mimetype='application/json',
        headers={
            'Content-Disposition': f'attachment; filename=legal-corpus-{source.lower()}.json',
            'Cache-Control': 'no-store',
        },
    )


@bp.route('/legal/corpus/import', methods=['POST'])
@login_required
def legal_corpus_import():
    if not current_user.can_manage_site():
        flash('Only the Website Controller can import legal corpus files.', 'danger')
        return redirect(url_for('legal.legal_home'))

    upload = request.files.get('corpus_file')
    source = (request.form.get('source') or 'ALL').strip().upper()
    if source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
        source = 'ALL'
    if not upload or not upload.filename:
        flash('Select a corpus file to import.', 'warning')
        return redirect(url_for('legal.legal_home'))

    filename = (upload.filename or '').strip().lower()
    if filename.endswith('.json'):
        try:
            payload = json.loads(upload.read().decode('utf-8'))
        except (UnicodeDecodeError, json.JSONDecodeError):
            flash('The uploaded JSON file is invalid.', 'danger')
            return redirect(url_for('legal.legal_home'))
    else:
        try:
            rows = _read_tabular_upload(upload)
        except ValueError:
            flash('Unsupported file type. Use .json, .csv, or .xlsx.', 'danger')
            return redirect(url_for('legal.legal_home'))
        except RuntimeError as exc:
            flash(str(exc), 'danger')
            return redirect(url_for('legal.legal_home'))
        except Exception:
            flash('Could not parse the uploaded corpus file.', 'danger')
            return redirect(url_for('legal.legal_home'))
        payload = _rows_to_payload(rows, source)

    result = import_corpus_payload(payload, source)
    flash(
        f"Imported {result['georgia_written']} Georgia entries, {result['ucmj_written']} UCMJ entries, and {result.get('base_order_written', 0)} Base Order entries.",
        'success',
    )
    return redirect(url_for('legal.legal_home'))
