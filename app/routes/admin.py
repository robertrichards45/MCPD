import csv
from collections import Counter
from datetime import datetime, timezone
import io
import json
from pathlib import Path
import re
import subprocess
import sys

from flask import Blueprint, Response, abort, flash, redirect, render_template, request, url_for
from flask_login import current_user, login_required
import os

from ..extensions import db
from ..models import AuditLog, StatCategory, StatsUpload
from ..permissions import can_access_builder_mode, can_manage_site
from ..services.excel_stats_parser import parse_targets
from ..services.legal_lookup import corpus_status, export_corpus_payload, get_entries, import_corpus_payload, reindex_corpus
from ..services.legal_ingestion import (
    fetch_official_uscode_source_markers,
    ingest_official_from_url,
    ingest_federal_official_from_url,
    ingest_uploaded_files,
)

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


bp = Blueprint('admin', __name__)
LEGAL_QA_STATUS_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'legal_qa_status.json'
LEGAL_QUERY_LOG_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'legal_query_log.jsonl'
LEGAL_BACKFILL_QUEUE_PATH = Path(__file__).resolve().parents[1] / 'data' / 'legal' / 'legal_backfill_queue.json'


def require_admin():
    if not can_manage_site(current_user):
        abort(403)


def require_builder_mode():
    if not can_access_builder_mode(current_user):
        abort(403)


def _load_legal_qa_status() -> dict | None:
    try:
        if not LEGAL_QA_STATUS_PATH.exists():
            return None
        payload = json.loads(LEGAL_QA_STATUS_PATH.read_text(encoding='utf-8'))
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def _write_legal_qa_status(payload: dict) -> None:
    try:
        LEGAL_QA_STATUS_PATH.parent.mkdir(parents=True, exist_ok=True)
        LEGAL_QA_STATUS_PATH.write_text(json.dumps(payload, indent=2), encoding='utf-8')
    except Exception:
        return


@bp.route('/admin/site-builder')
@login_required
def site_builder():
    require_builder_mode()
    return render_template('site_builder.html', title='Site Builder', user=current_user)


def _load_legal_query_events(limit: int = 5000) -> list[dict]:
    if not LEGAL_QUERY_LOG_PATH.exists():
        return []
    lines = []
    try:
        with LEGAL_QUERY_LOG_PATH.open('r', encoding='utf-8') as handle:
            for line in handle:
                line = line.strip()
                if not line:
                    continue
                lines.append(line)
    except Exception:
        return []
    events = []
    for line in lines[-limit:]:
        try:
            payload = json.loads(line)
            if isinstance(payload, dict):
                events.append(payload)
        except Exception:
            continue
    return events


def _clear_legal_query_events() -> None:
    try:
        if LEGAL_QUERY_LOG_PATH.exists():
            LEGAL_QUERY_LOG_PATH.unlink()
    except Exception:
        return


def _load_backfill_queue() -> list[dict]:
    try:
        if not LEGAL_BACKFILL_QUEUE_PATH.exists():
            return []
        payload = json.loads(LEGAL_BACKFILL_QUEUE_PATH.read_text(encoding='utf-8'))
        if not isinstance(payload, list):
            return []
        clean: list[dict] = []
        for item in payload:
            if not isinstance(item, dict):
                continue
            clean.append(item)
        return clean
    except Exception:
        return []


def _write_backfill_queue(rows: list[dict]) -> None:
    try:
        LEGAL_BACKFILL_QUEUE_PATH.parent.mkdir(parents=True, exist_ok=True)
        LEGAL_BACKFILL_QUEUE_PATH.write_text(json.dumps(rows, indent=2), encoding='utf-8')
    except Exception:
        return


def _suggest_source_for_term(term: str) -> str:
    t = (term or '').lower()
    if any(k in t for k in ('article', 'awol', 'barracks', 'marine', 'ucmj', 'lawful order', 'insubordinate')):
        return 'UCMJ'
    if any(k in t for k in ('usc', 'federal', 'interstate', 'counterfeit', 'identity', 'bank', 'mail', 'wire')):
        return 'FEDERAL_USC'
    if any(k in t for k in ('mclb', 'base order', 'installation', 'albany order')):
        return 'BASE_ORDER'
    return 'GEORGIA'


def _normalize_source_label(value: str, fallback: str = 'ALL') -> str:
    raw = (value or '').strip().upper()
    if raw in {'GA', 'GEORGIA', 'GEORGIA_CODE'}:
        return 'GEORGIA'
    if raw in {'UCMJ', 'MCM', 'ARTICLE'}:
        return 'UCMJ'
    if raw in {'BASE_ORDER', 'BASE', 'ORDER', 'MCLBAO'}:
        return 'BASE_ORDER'
    if raw in {'FEDERAL_USC', 'FEDERAL', 'USC', 'UNITED_STATES_CODE'}:
        return 'FEDERAL_USC'
    if raw in {'ALL', ''}:
        return fallback
    return fallback


def _split_multi(value: str) -> list[str]:
    text = (value or '').strip()
    if not text:
        return []
    for sep in ('|', ';', ','):
        if sep in text:
            return [item.strip() for item in text.split(sep) if item.strip()]
    return [text]


def _split_multi_any(value: str) -> list[str]:
    text = (value or '').strip()
    if not text:
        return []
    parts = re.split(r'[|;,]', text)
    return [item.strip() for item in parts if item.strip()]


def _rows_to_payload(rows: list[dict], target_source: str) -> dict:
    georgia_entries = []
    ucmj_entries = []
    base_order_entries = []
    federal_usc_entries = []
    for row in rows:
        code = (row.get('code') or '').strip()
        title = (row.get('title') or '').strip()
        summary = (row.get('summary') or '').strip()
        if not (code and title and summary):
            continue
        row_source = _normalize_source_label(row.get('source', ''), target_source if target_source != 'ALL' else 'GEORGIA')
        if target_source in {'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
            row_source = target_source
        entry = {
            'source': row_source,
            'code': code,
            'title': title,
            'summary': summary,
            'elements': _split_multi(row.get('elements', '')),
            'notes': (row.get('notes') or '').strip(),
            'keywords': _split_multi(row.get('keywords', '')),
            'related_codes': _split_multi(row.get('related_codes', '')),
            'minimum_punishment': (row.get('minimum_punishment') or '').strip(),
            'maximum_punishment': (row.get('maximum_punishment') or '').strip(),
        }
        if row_source == 'UCMJ':
            ucmj_entries.append(entry)
        elif row_source == 'BASE_ORDER':
            base_order_entries.append(entry)
        elif row_source == 'FEDERAL_USC':
            federal_usc_entries.append(entry)
        else:
            georgia_entries.append(entry)
    return {
        'georgia_codes': georgia_entries,
        'ucmj_articles': ucmj_entries,
        'base_orders': base_order_entries,
        'federal_usc_codes': federal_usc_entries,
    }


def _read_tabular_upload(upload) -> list[dict]:
    filename = (upload.filename or '').strip()
    ext = Path(filename).suffix.lower()
    if ext == '.csv':
        text = upload.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        return [{(k or '').strip().lower(): (v or '').strip() for k, v in row.items()} for row in reader]
    if ext in {'.xlsx', '.xlsm'}:
        if load_workbook is None:
            raise RuntimeError('openpyxl is required for Excel imports')
        upload.seek(0)
        workbook = load_workbook(upload, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [(str(cell or '').strip().lower()) for cell in rows[0]]
        out: list[dict] = []
        for row in rows[1:]:
            record = {}
            for i, key in enumerate(headers):
                if not key:
                    continue
                value = row[i] if i < len(row) else ''
                record[key] = '' if value is None else str(value).strip()
            if any(record.values()):
                out.append(record)
        return out
    raise ValueError('Unsupported file type')


@bp.route('/admin/stats/categories', methods=['GET', 'POST'])
@login_required
def stats_categories():
    require_admin()
    if request.method == 'POST':
        for key, val in request.form.items():
            if key.startswith('target_'):
                cat_id = int(key.split('_')[1])
                cat = StatCategory.query.get(cat_id)
                if cat:
                    try:
                        cat.target_value = int(val or 0)
                    except ValueError:
                        cat.target_value = 0
        db.session.commit()
        return redirect(url_for('admin.stats_categories'))

    categories = StatCategory.query.order_by(StatCategory.name).all()
    return render_template('stats_targets.html', categories=categories, user=current_user)


@bp.route('/admin/stats/targets-upload', methods=['GET', 'POST'])
@login_required
def stats_targets_upload():
    require_admin()
    if request.method == 'POST':
        file = request.files.get('file')
        if not file:
            return render_template('stats_targets_upload.html', error='No file uploaded.', user=current_user)
        rows, layout = parse_targets(file)
        updated = 0
        for row in rows:
            name = row['category']
            target = row['target']
            category = StatCategory.query.filter_by(name=name).first()
            if not category:
                category = StatCategory(name=name, target_value=target)
                db.session.add(category)
            else:
                category.target_value = target
            updated += 1
        db.session.add(AuditLog(actor_id=current_user.id, action='stats_targets_upload', details=f'{updated} targets ({layout})'))
        db.session.commit()
        return redirect(url_for('admin.stats_categories'))

    return render_template('stats_targets_upload.html', user=current_user)


@bp.route('/admin/stats/uploads')
@login_required
def stats_uploads():
    require_admin()
    uploads = StatsUpload.query.order_by(StatsUpload.uploaded_at.desc()).all()
    items = []
    for upload in uploads:
        summary = {}
        if upload.parse_summary_json:
            try:
                summary = json.loads(upload.parse_summary_json)
            except Exception:
                summary = {}
        items.append({'upload': upload, 'summary': summary})
    return render_template('stats_uploads.html', items=items, user=current_user)


@bp.route('/admin/stats/uploads/<int:upload_id>')
@login_required
def stats_upload_detail(upload_id):
    require_admin()
    upload = StatsUpload.query.get_or_404(upload_id)
    summary = {}
    if upload.parse_summary_json:
        try:
            summary = json.loads(upload.parse_summary_json)
        except Exception:
            summary = {}
    return render_template('stats_upload_detail.html', upload=upload, summary=summary, user=current_user)


@bp.route('/admin/legal-corpus', methods=['GET', 'POST'])
@login_required
def legal_corpus_admin():
    require_admin()
    qa_result = _load_legal_qa_status()
    source_markers = fetch_official_uscode_source_markers(timeout_sec=4)
    if request.method == 'POST':
        action = (request.form.get('action') or '').strip().lower()
        if action == 'run_qa':
            repo_root = Path(__file__).resolve().parents[2]
            script_path = repo_root / 'app' / 'scripts' / 'legal_regression_check.py'
            try:
                run = subprocess.run(
                    [sys.executable, str(script_path)],
                    cwd=str(repo_root),
                    capture_output=True,
                    text=True,
                    timeout=60,
                    check=False,
                )
                output = (run.stdout or '').strip()
                error = (run.stderr or '').strip()
                qa_result = {
                    'ok': run.returncode == 0,
                    'code': run.returncode,
                    'output': output.splitlines() if output else [],
                    'error': error.splitlines() if error else [],
                    'ran_at_utc': datetime.now(timezone.utc).isoformat(),
                }
                _write_legal_qa_status(qa_result)
                db.session.add(
                    AuditLog(
                        actor_id=current_user.id,
                        action='legal_qa_run',
                        details=f"exit_code={run.returncode}",
                    )
                )
                db.session.commit()
            except Exception as exc:
                qa_result = {
                    'ok': False,
                    'code': -1,
                    'output': [],
                    'error': [f'QA execution failed: {exc}'],
                    'ran_at_utc': datetime.now(timezone.utc).isoformat(),
                }
                _write_legal_qa_status(qa_result)
            return render_template(
                'admin_legal_corpus.html',
                user=current_user,
                corpus_status=corpus_status(),
                qa_result=qa_result,
                source_markers=source_markers,
            )
        if action == 'reindex':
            status = reindex_corpus()
            flash(
                (
                    "Legal corpus reindexed. "
                    f"Georgia={status['georgia_count']}, UCMJ={status['ucmj_count']}, Base Orders={status.get('base_order_count', 0)}, Federal USC={status.get('federal_usc_count', 0)}"
                ),
                'success',
            )
            db.session.add(
                AuditLog(
                    actor_id=current_user.id,
                    action='legal_reindex',
                    details='manual corpus reindex from admin',
                )
            )
            db.session.commit()
            return redirect(url_for('admin.legal_corpus_admin'))
        if action == 'ingest_sources':
            source = (request.form.get('source') or 'ALL').strip().upper()
            if source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
                source = 'ALL'
            use_ai = (request.form.get('use_ai_enrichment') or '').strip().lower() in {'1', 'true', 'on', 'yes'}
            uploads = request.files.getlist('source_files')
            if not uploads:
                flash('Select one or more source files to ingest.', 'warning')
                return redirect(url_for('admin.legal_corpus_admin'))
            ingest_result = ingest_uploaded_files(uploads, source=source, use_ai_enrichment=use_ai)
            if not ingest_result.entries:
                warning_text = '; '.join(ingest_result.warnings) if ingest_result.warnings else 'No parsable legal entries found.'
                flash(f'Ingestion completed with no entries imported. {warning_text}', 'warning')
                return redirect(url_for('admin.legal_corpus_admin'))

            grouped_payload = {
                'georgia_codes': [entry for entry in ingest_result.entries if entry.get('source') == 'GEORGIA'],
                'ucmj_articles': [entry for entry in ingest_result.entries if entry.get('source') == 'UCMJ'],
                'base_orders': [entry for entry in ingest_result.entries if entry.get('source') == 'BASE_ORDER'],
                'federal_usc_codes': [entry for entry in ingest_result.entries if entry.get('source') == 'FEDERAL_USC'],
            }
            result = import_corpus_payload(grouped_payload, 'ALL')
            warning_text = '; '.join(ingest_result.warnings[:3])
            flash(
                (
                    f"Ingested {ingest_result.entries_built} entries from {ingest_result.files_processed} files. "
                    f"Imported Georgia={result['georgia_written']}, UCMJ={result['ucmj_written']}, Base Orders={result.get('base_order_written', 0)}, Federal USC={result.get('federal_usc_written', 0)}."
                    + (f" Warnings: {warning_text}" if warning_text else '')
                ),
                'success',
            )
            db.session.add(
                AuditLog(
                    actor_id=current_user.id,
                    action='legal_source_ingestion',
                    details=(
                        f"source={source} files={ingest_result.files_processed} built={ingest_result.entries_built} "
                        f"use_ai={use_ai} warnings={len(ingest_result.warnings)}"
                    ),
                )
            )
            db.session.commit()
            return redirect(url_for('admin.legal_corpus_admin'))
        if action == 'ingest_federal_url':
            source_url = (request.form.get('source_url') or '').strip()
            ingest_result = ingest_federal_official_from_url(source_url, timeout_sec=20)
            if not ingest_result.entries:
                warning_text = '; '.join(ingest_result.warnings) if ingest_result.warnings else 'No federal entries found.'
                flash(f'Federal URL ingestion finished with no entries. {warning_text}', 'warning')
                return redirect(url_for('admin.legal_corpus_admin'))
            grouped_payload = {
                'federal_usc_codes': list(ingest_result.entries),
            }
            result = import_corpus_payload(grouped_payload, 'ALL')
            warning_text = '; '.join(ingest_result.warnings[:3])
            flash(
                (
                    f"Federal URL ingestion imported {result.get('federal_usc_written', 0)} records."
                    + (f" Warnings: {warning_text}" if warning_text else '')
                ),
                'success',
            )
            db.session.add(
                AuditLog(
                    actor_id=current_user.id,
                    action='legal_federal_url_ingest',
                    details=f"url={source_url} imported={result.get('federal_usc_written', 0)} warnings={len(ingest_result.warnings)}",
                )
            )
            db.session.commit()
            return redirect(url_for('admin.legal_corpus_admin'))
        if action == 'ingest_source_url':
            source_url = (request.form.get('source_url') or '').strip()
            source = _normalize_source_label((request.form.get('source') or 'ALL').strip(), fallback='ALL')
            use_ai = bool((request.form.get('use_ai_enrichment') or '').strip())
            ingest_result = ingest_official_from_url(source_url, source=source, timeout_sec=20, use_ai_enrichment=use_ai)
            if not ingest_result.entries:
                warning_text = '; '.join(ingest_result.warnings) if ingest_result.warnings else 'No importable entries found.'
                flash(f'URL ingestion finished with no entries. {warning_text}', 'warning')
                return redirect(url_for('admin.legal_corpus_admin'))
            grouped_payload: dict[str, list[dict]] = {
                'georgia_codes': [],
                'ucmj_articles': [],
                'base_orders': [],
                'federal_usc_codes': [],
            }
            for row in ingest_result.entries:
                src = str(row.get('source') or '').upper()
                if src == 'GEORGIA':
                    grouped_payload['georgia_codes'].append(dict(row))
                elif src == 'UCMJ':
                    grouped_payload['ucmj_articles'].append(dict(row))
                elif src == 'BASE_ORDER':
                    grouped_payload['base_orders'].append(dict(row))
                elif src == 'FEDERAL_USC':
                    grouped_payload['federal_usc_codes'].append(dict(row))
            result = import_corpus_payload(grouped_payload, 'ALL')
            warning_text = '; '.join(ingest_result.warnings[:3])
            flash(
                (
                    f"URL ingestion imported GEORGIA={result.get('georgia_written', 0)}, "
                    f"UCMJ={result.get('ucmj_written', 0)}, BASE_ORDER={result.get('base_order_written', 0)}, "
                    f"FEDERAL={result.get('federal_usc_written', 0)}."
                    + (f" Warnings: {warning_text}" if warning_text else '')
                ),
                'success',
            )
            db.session.add(
                AuditLog(
                    actor_id=current_user.id,
                    action='legal_source_url_ingest',
                    details=(
                        f"source={source} url={source_url} use_ai={use_ai} "
                        f"imported={sum(result.get(k, 0) for k in ('georgia_written','ucmj_written','base_order_written','federal_usc_written'))} "
                        f"warnings={len(ingest_result.warnings)}"
                    ),
                )
            )
            db.session.commit()
            return redirect(url_for('admin.legal_corpus_admin'))

        upload = request.files.get('corpus_file')
        source = (request.form.get('source') or 'ALL').strip().upper()
        if source not in {'ALL', 'GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'}:
            source = 'ALL'
        if not upload or not upload.filename:
            flash('Select a corpus file to import.', 'warning')
            return redirect(url_for('admin.legal_corpus_admin'))

        filename = (upload.filename or '').strip().lower()
        if filename.endswith('.json'):
            try:
                payload = json.loads(upload.read().decode('utf-8'))
            except (UnicodeDecodeError, json.JSONDecodeError):
                flash('The uploaded JSON file is invalid.', 'danger')
                return redirect(url_for('admin.legal_corpus_admin'))
        else:
            try:
                rows = _read_tabular_upload(upload)
            except ValueError:
                flash('Unsupported file type. Use .json, .csv, or .xlsx.', 'danger')
                return redirect(url_for('admin.legal_corpus_admin'))
            except RuntimeError as exc:
                flash(str(exc), 'danger')
                return redirect(url_for('admin.legal_corpus_admin'))
            except Exception:
                flash('Could not parse the uploaded corpus file.', 'danger')
                return redirect(url_for('admin.legal_corpus_admin'))
            payload = _rows_to_payload(rows, source)

        result = import_corpus_payload(payload, source)
        flash(
            f"Imported {result['georgia_written']} Georgia entries, {result['ucmj_written']} UCMJ entries, {result.get('base_order_written', 0)} Base Order entries, and {result.get('federal_usc_written', 0)} Federal USC entries.",
            'success',
        )
        db.session.add(
            AuditLog(
                actor_id=current_user.id,
                action='legal_corpus_import',
                details=f"source={source} georgia={result['georgia_written']} ucmj={result['ucmj_written']} base_order={result.get('base_order_written', 0)} federal_usc={result.get('federal_usc_written', 0)}",
            )
        )
        db.session.commit()
        return redirect(url_for('admin.legal_corpus_admin'))

    return render_template(
        'admin_legal_corpus.html',
        user=current_user,
        corpus_status=corpus_status(),
        qa_result=qa_result,
        source_markers=source_markers,
    )


@bp.route('/admin/legal-analytics')
@login_required
def legal_analytics():
    require_admin()
    events = _load_legal_query_events()
    query_counter = Counter()
    zero_counter = Counter()
    low_counter = Counter()
    code_counter = Counter()
    source_counter = Counter()
    for event in events:
        query = str(event.get('query') or '').strip()
        if not query:
            continue
        source = str(event.get('source') or 'ALL').strip().upper()
        count = int(event.get('result_count') or 0)
        query_counter[query] += 1
        source_counter[source] += 1
        if count == 0:
            zero_counter[query] += 1
        elif count <= 1:
            low_counter[query] += 1
        for code in event.get('top_codes') or []:
            code_counter[str(code)] += 1

    status = corpus_status()
    coverage = status.get('coverage') if isinstance(status.get('coverage'), dict) else {}

    known_terms: set[str] = set()
    for source in ('GEORGIA', 'UCMJ', 'BASE_ORDER', 'FEDERAL_USC'):
        for entry in get_entries(source):
            tokens = re.findall(r'[a-z0-9]+', ' '.join([
                entry.code,
                entry.title,
                entry.summary,
                ' '.join(entry.keywords),
                ' '.join(entry.aliases),
                ' '.join(entry.synonyms),
                ' '.join(entry.narrative_triggers),
                ' '.join(entry.examples),
            ]).lower())
            for token in tokens:
                if len(token) >= 4:
                    known_terms.add(token)

    weak_phrase_counter = Counter()
    missing_term_counter = Counter()
    for query, count in low_counter.items():
        weak_phrase_counter[query] += count
    for query, count in zero_counter.items():
        weak_phrase_counter[query] += count
    for query, count in weak_phrase_counter.items():
        parts = re.findall(r'[a-z0-9]+', query.lower())
        for part in parts:
            if len(part) < 4:
                continue
            if part not in known_terms:
                missing_term_counter[part] += count

    likely_coverage_gaps = []
    if zero_counter:
        likely_coverage_gaps.append(('No-result narratives', sum(zero_counter.values())))
    if missing_term_counter:
        likely_coverage_gaps.append(('Unmapped officer terms', sum(missing_term_counter.values())))
    for src, metrics in coverage.items():
        if not isinstance(metrics, dict):
            continue
        gap_score = (
            int(metrics.get('missing_aliases') or 0)
            + int(metrics.get('missing_examples') or 0)
            + int(metrics.get('missing_punishment') or 0)
            + int(metrics.get('missing_official_text') or 0)
        )
        likely_coverage_gaps.append((f'{src} structural gaps', gap_score))
    likely_coverage_gaps = sorted(likely_coverage_gaps, key=lambda item: item[1], reverse=True)[:12]

    weak_term_suggestions = [
        (term, count, _suggest_source_for_term(term))
        for term, count in missing_term_counter.most_common(50)
    ]
    backfill_queue = _load_backfill_queue()

    return render_template(
        'admin_legal_analytics.html',
        user=current_user,
        total_events=len(events),
        top_queries=query_counter.most_common(30),
        zero_queries=zero_counter.most_common(30),
        low_queries=low_counter.most_common(30),
        top_codes=code_counter.most_common(30),
        source_counts=source_counter.most_common(),
        coverage=coverage,
        coverage_warnings=status.get('coverage_warnings') or [],
        weak_term_suggestions=weak_term_suggestions,
        likely_coverage_gaps=likely_coverage_gaps,
        backfill_queue=backfill_queue[:100],
        backfill_pending=sum(1 for item in backfill_queue if str(item.get('status') or 'pending') == 'pending'),
    )


@bp.route('/admin/legal-analytics/export.csv')
@login_required
def legal_analytics_export_csv():
    require_admin()
    events = _load_legal_query_events(limit=50000)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ts_utc', 'query', 'source', 'result_count', 'top_confidence', 'weak_result', 'top_codes'])
    for event in events:
        writer.writerow([
            event.get('ts_utc', ''),
            event.get('query', ''),
            event.get('source', ''),
            event.get('result_count', ''),
            event.get('top_confidence', ''),
            event.get('weak_result', ''),
            ' | '.join(event.get('top_codes') or []),
        ])
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=legal-analytics.csv'},
    )


@bp.route('/admin/legal-analytics/clear', methods=['POST'])
@login_required
def legal_analytics_clear():
    require_admin()
    _clear_legal_query_events()
    db.session.add(
        AuditLog(
            actor_id=current_user.id,
            action='legal_analytics_clear',
            details='cleared legal query telemetry log',
        )
    )
    db.session.commit()
    flash('Legal analytics log cleared.', 'success')
    return redirect(url_for('admin.legal_analytics'))


@bp.route('/admin/legal-analytics/backfill-generate', methods=['POST'])
@login_required
def legal_analytics_backfill_generate():
    require_admin()
    events = _load_legal_query_events(limit=50000)
    term_counter = Counter()
    phrase_counter = Counter()
    for event in events:
        query = str(event.get('query') or '').strip().lower()
        if not query:
            continue
        result_count = int(event.get('result_count') or 0)
        top_conf = int(event.get('top_confidence') or 0)
        if result_count == 0 or top_conf < 55:
            phrase_counter[query] += 1
            for token in re.findall(r'[a-z0-9]+', query):
                if len(token) >= 4:
                    term_counter[token] += 1

    rows: list[dict] = []
    for term, count in term_counter.most_common(100):
        rows.append(
            {
                'type': 'term',
                'value': term,
                'suggested_source': _suggest_source_for_term(term),
                'occurrences': count,
                'reason': 'High-frequency unmapped/weak term',
                'status': 'pending',
                'created_at_utc': datetime.now(timezone.utc).isoformat(),
            }
        )
    for phrase, count in phrase_counter.most_common(80):
        rows.append(
            {
                'type': 'phrase',
                'value': phrase,
                'suggested_source': _suggest_source_for_term(phrase),
                'occurrences': count,
                'reason': 'Low-confidence or zero-result officer narrative',
                'status': 'pending',
                'created_at_utc': datetime.now(timezone.utc).isoformat(),
            }
        )
    _write_backfill_queue(rows)
    db.session.add(
        AuditLog(
            actor_id=current_user.id,
            action='legal_backfill_generate',
            details=f'rows={len(rows)} terms={min(100, len(term_counter))} phrases={min(80, len(phrase_counter))}',
        )
    )
    db.session.commit()
    flash(f'Generated backfill queue with {len(rows)} items.', 'success')
    return redirect(url_for('admin.legal_analytics'))


@bp.route('/admin/legal-analytics/backfill-clear', methods=['POST'])
@login_required
def legal_analytics_backfill_clear():
    require_admin()
    _write_backfill_queue([])
    db.session.add(
        AuditLog(
            actor_id=current_user.id,
            action='legal_backfill_clear',
            details='cleared legal backfill queue',
        )
    )
    db.session.commit()
    flash('Backfill queue cleared.', 'success')
    return redirect(url_for('admin.legal_analytics'))


def _looks_federal_query(query: str) -> bool:
    text = (query or '').lower()
    return bool(re.search(
        r'federal|usc|united states code|interstate|bank robbery|wire fraud|identity theft|mail theft|government property|federal facility|government computer|counterfeit',
        text,
    ))


@bp.route('/admin/legal-federal-review', methods=['GET', 'POST'])
@login_required
def legal_federal_review():
    require_admin()

    if request.method == 'POST':
        action = (request.form.get('action') or '').strip().lower()
        code = (request.form.get('code') or '').strip()
        payload_all = export_corpus_payload('ALL')
        federal_rows = list(payload_all.get('federal_usc_codes') or [])
        modified = False
        changed_code = ''
        for row in federal_rows:
            if str(row.get('code') or '').strip() != code:
                continue
            changed_code = code
            if action == 'toggle_active':
                row['active_flag'] = not bool(row.get('active_flag', True))
                row['last_updated'] = datetime.now(timezone.utc).isoformat()
                modified = True
            elif action == 'save_record':
                row['category'] = (request.form.get('category') or '').strip()
                row['subcategory'] = (request.form.get('subcategory') or '').strip()
                row['severity'] = (request.form.get('severity') or '').strip()
                row['source_reference_url'] = (request.form.get('source_reference_url') or '').strip()
                row['official_text'] = (request.form.get('official_text') or '').strip()
                row['official_text_available'] = bool((request.form.get('official_text_available') or '').strip().lower() in {'1', 'true', 'on', 'yes'})
                row['citation_requires_verification'] = bool((request.form.get('citation_requires_verification') or '').strip().lower() in {'1', 'true', 'on', 'yes'})
                row['aliases'] = _split_multi_any(request.form.get('aliases') or '')
                row['synonyms'] = _split_multi_any(request.form.get('synonyms') or '')
                row['narrative_triggers'] = _split_multi_any(request.form.get('narrative_triggers') or '')
                row['derived_aliases'] = _split_multi_any(request.form.get('derived_aliases') or '')
                row['derived_synonyms'] = _split_multi_any(request.form.get('derived_synonyms') or '')
                row['derived_triggers'] = _split_multi_any(request.form.get('derived_triggers') or '')
                row['derived_examples'] = _split_multi_any(request.form.get('derived_examples') or '')
                row['derived_summary'] = (request.form.get('derived_summary') or '').strip()
                row['minimum_punishment'] = (request.form.get('minimum_punishment') or '').strip()
                row['maximum_punishment'] = (request.form.get('maximum_punishment') or '').strip()
                row['last_updated'] = datetime.now(timezone.utc).isoformat()
                modified = True
            break

        if modified:
            payload_all['federal_usc_codes'] = federal_rows
            result = import_corpus_payload(payload_all, 'ALL')
            db.session.add(
                AuditLog(
                    actor_id=current_user.id,
                    action='legal_federal_review_update',
                    details=f"action={action} code={changed_code} federal_written={result.get('federal_usc_written', 0)}",
                )
            )
            db.session.commit()
            flash(f'Federal record updated: {changed_code}', 'success')
        else:
            flash('No changes applied.', 'warning')
        return redirect(url_for('admin.legal_federal_review', code=code))

    payload_all = export_corpus_payload('ALL')
    federal_rows = list(payload_all.get('federal_usc_codes') or [])
    federal_rows.sort(key=lambda item: str(item.get('code') or ''))
    selected_code = (request.args.get('code') or '').strip()
    selected = None
    for row in federal_rows:
        if str(row.get('code') or '').strip() == selected_code:
            selected = row
            break
    if selected is None and federal_rows:
        selected = federal_rows[0]
        selected_code = str(selected.get('code') or '')

    events = _load_legal_query_events(limit=10000)
    weak_federal_counter = Counter()
    for event in events:
        query = str(event.get('query') or '').strip()
        if not query:
            continue
        source = str(event.get('source') or '').strip().upper()
        result_count = int(event.get('result_count') or 0)
        top_codes = [str(code) for code in (event.get('top_codes') or [])]
        has_federal_code = any(code.upper().startswith(('18 USC', '21 USC', '26 USC', '49 USC')) for code in top_codes)
        if result_count <= 1 and (source == 'FEDERAL_USC' or _looks_federal_query(query) or has_federal_code):
            weak_federal_counter[query] += 1

    summary = {
        'total': len(federal_rows),
        'active': sum(1 for row in federal_rows if bool(row.get('active_flag', True))),
        'missing_official_text': sum(1 for row in federal_rows if not bool(row.get('official_text_available', False))),
        'citation_verify': sum(1 for row in federal_rows if bool(row.get('citation_requires_verification', False))),
        'missing_punishment': sum(
            1 for row in federal_rows
            if not ((row.get('minimum_punishment') or '').strip() and (row.get('maximum_punishment') or '').strip())
        ),
    }

    return render_template(
        'admin_legal_federal_review.html',
        user=current_user,
        corpus_status=corpus_status(),
        federal_rows=federal_rows,
        selected=selected,
        selected_code=selected_code,
        summary=summary,
        weak_federal_queries=weak_federal_counter.most_common(40),
    )


@bp.route('/admin/system-status')
@login_required
def system_status():
    require_admin()
    api_key_raw = os.environ.get('OPENAI_API_KEY', '')
    ai_key_set = bool(api_key_raw.strip())
    ai_key_preview = f'{api_key_raw[:8]}…' if ai_key_set else '(not set)'
    from ..services.ai_client import _AI_DISABLED_MESSAGE, _AI_DISABLED_UNTIL
    from datetime import datetime, timezone as tz
    ai_cooldown_active = bool(
        _AI_DISABLED_MESSAGE
        and _AI_DISABLED_UNTIL
        and datetime.now(tz.utc) < _AI_DISABLED_UNTIL
    )
    return render_template(
        'admin_system_status.html',
        user=current_user,
        ai_key_set=ai_key_set,
        ai_key_preview=ai_key_preview,
        ai_cooldown_active=ai_cooldown_active,
        ai_cooldown_message=_AI_DISABLED_MESSAGE if ai_cooldown_active else '',
        railway_env=os.environ.get('RAILWAY_ENVIRONMENT', ''),
    )
