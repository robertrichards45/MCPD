import json
from datetime import datetime, timezone
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def _require_openpyxl():
    if load_workbook is None:
        raise RuntimeError('openpyxl is not installed')


def detect_layout(ws):
    headers = [c.value for c in ws[1]]
    if headers and len(headers) >= 3 and 'Officer' in headers and 'Category' in headers and 'Value' in headers:
        return 'A'
    return 'B'


def parse_excel(path):
    _require_openpyxl()
    wb = load_workbook(path)
    ws = wb.active
    layout = detect_layout(ws)
    rows = []

    if layout == 'A':
        header = [c.value for c in ws[1]]
        col_map = {name: idx for idx, name in enumerate(header)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            officer = row[col_map['Officer']]
            category = row[col_map['Category']]
            value = row[col_map['Value']]
            rows.append({'officer': officer, 'category': category, 'value': int(value or 0)})
    else:
        header = [c.value for c in ws[1]]
        categories = header[1:]
        for row in ws.iter_rows(min_row=2, values_only=True):
            officer = row[0]
            if not officer:
                continue
            for idx, cat in enumerate(categories, start=1):
                value = row[idx] if idx < len(row) else 0
                rows.append({'officer': officer, 'category': cat, 'value': int(value or 0)})

    return rows, layout


def parse_targets(file_like):
    _require_openpyxl()
    wb = load_workbook(file_like)
    ws = wb.active
    headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]

    # Layout T1: Category, Target
    if 'Category' in headers and any(h in headers for h in ['Target', 'TargetValue', 'Target Value']):
        cat_idx = headers.index('Category')
        if 'Target' in headers:
            tgt_idx = headers.index('Target')
        elif 'TargetValue' in headers:
            tgt_idx = headers.index('TargetValue')
        else:
            tgt_idx = headers.index('Target Value')
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            category = row[cat_idx]
            target = row[tgt_idx]
            if category:
                rows.append({'category': str(category).strip(), 'target': int(target or 0)})
        return rows, 'T1'

    # Layout T2: First row categories, second row targets, first cell is TARGETS
    if len(headers) > 1 and ws.max_row >= 2:
        first_cell = str(ws[2][0].value).strip() if ws[2][0].value is not None else ''
        if first_cell.upper() in ['TARGET', 'TARGETS']:
            categories = headers[1:]
            target_row = [c.value for c in ws[2]][1:]
            rows = []
            for cat, val in zip(categories, target_row):
                if cat:
                    rows.append({'category': str(cat).strip(), 'target': int(val or 0)})
            return rows, 'T2'

    return [], 'UNKNOWN'


def summary(rows):
    return json.dumps({
        'rows': len(rows),
        'categories': len(set(r['category'] for r in rows)),
        'generated_at': datetime.now(timezone.utc).replace(tzinfo=None).isoformat()
    })
