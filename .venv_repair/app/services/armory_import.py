import csv
import io

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


EXPECTED_COLUMNS = {
    'asset_type',
    'label',
    'serial_number',
    'rack_number',
    'radio_identifier',
    'oc_identifier',
}


def _normalized(value):
    return (value or '').strip()


def _normalize_row(row):
    return {
        'asset_type': _normalized(row.get('asset_type')).upper(),
        'label': _normalized(row.get('label')),
        'serial_number': _normalized(row.get('serial_number')).upper(),
        'rack_number': _normalized(row.get('rack_number')) or None,
        'radio_identifier': _normalized(row.get('radio_identifier')) or None,
        'oc_identifier': _normalized(row.get('oc_identifier')) or None,
    }


def _parse_csv(stream):
    text = stream.read()
    if isinstance(text, bytes):
        text = text.decode('utf-8-sig', errors='ignore')
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append(_normalize_row(row))
    return rows


def _parse_xlsx(stream):
    if load_workbook is None:
        raise ValueError('XLSX parsing requires openpyxl. Install openpyxl or upload CSV.')
    data = stream.read()
    workbook = load_workbook(io.BytesIO(data), data_only=True)
    sheet = workbook.active
    header_cells = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or []
    headers = [str(value).strip() if value is not None else '' for value in header_cells]
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None and str(value).strip() for value in values):
            continue
        row_map = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            row_map[header] = values[index] if index < len(values) else ''
        rows.append(_normalize_row(row_map))
    return rows


def parse_armory_asset_upload(upload_file):
    filename = (getattr(upload_file, 'filename', '') or '').lower()
    if filename.endswith('.csv'):
        rows = _parse_csv(upload_file.stream)
    elif filename.endswith('.xlsx'):
        rows = _parse_xlsx(upload_file.stream)
    else:
        raise ValueError('Only CSV and XLSX files are supported for armory import.')

    cleaned = []
    for row in rows:
        if not row['asset_type'] and not row['label'] and not row['serial_number']:
            continue
        if not row['asset_type'] or not row['label'] or not row['serial_number']:
            raise ValueError('Each import row must include asset_type, label, and serial_number.')
        cleaned.append(row)
    return cleaned
