import os
import shutil
from datetime import datetime

from flask import current_app
try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover
    Workbook = None
    load_workbook = None

from ..models import TruckGateLog
from .truck_gate_import import DEFAULT_TRUCK_GATE_SOURCE, TRUCK_GATE_SOURCE_SHEET


TRUCK_GATE_EXPORT_HEADERS = [
    'DRIVER NAME DRIVER ',
    'LICENCE NUMBER AND STATE',
    'COMPANY NAME',
    'PHONE NUMBER',
    'TYPE',
    'VEHICLE REGISTRATION, STATE',
    'MAKE MODEL COLOR',
    'DESTINATION',
    'TYPE ',
]

TRUCK_GATE_DATA_START_ROW = 11


def _daily_folder_parts(log_date):
    dt = datetime.strptime(log_date, '%Y-%m-%d')
    return [dt.strftime('%Y'), dt.strftime('%B'), log_date]


def daily_workbook_filename(log_date):
    return f'truck-gate-{log_date}.xlsx'


def daily_workbook_relpath(log_date):
    return os.path.join('generated', 'truck_gate_logs', *_daily_folder_parts(log_date), daily_workbook_filename(log_date))


def daily_workbook_abspath(log_date):
    return os.path.join(current_app.root_path, daily_workbook_relpath(log_date))


def ensure_daily_workbook_directory(log_date):
    folder = os.path.dirname(daily_workbook_abspath(log_date))
    os.makedirs(folder, exist_ok=True)
    return folder


def _write_database_sheet(workbook, logs):
    if TRUCK_GATE_SOURCE_SHEET in workbook.sheetnames:
        sheet = workbook[TRUCK_GATE_SOURCE_SHEET]
    else:
        sheet = workbook.active
        sheet.title = TRUCK_GATE_SOURCE_SHEET

    for column_index, header in enumerate(TRUCK_GATE_EXPORT_HEADERS, start=1):
        sheet.cell(row=1, column=column_index, value=header)

    max_clear_row = max(sheet.max_row, TRUCK_GATE_DATA_START_ROW + len(logs) + 50)
    for row_index in range(TRUCK_GATE_DATA_START_ROW, max_clear_row + 1):
        for column_index in range(1, len(TRUCK_GATE_EXPORT_HEADERS) + 1):
            sheet.cell(row=row_index, column=column_index, value=None)

    for row_index, log in enumerate(logs, start=TRUCK_GATE_DATA_START_ROW):
        driver = log.driver
        company = log.company or (driver.company if driver else None)
        vehicle = log.vehicle
        license_display = ''
        plate_display = ''
        phone_number = ''
        vehicle_type = ''
        destination = ''
        if driver:
            license_display = ' '.join(part for part in [driver.license_number or '', driver.license_state or ''] if part).strip()
            phone_number = driver.phone_number or ''
            vehicle_type = driver.vehicle_type or ''
            destination = driver.destination or ''
        if not phone_number and company:
            phone_number = company.phone_number or ''
        if vehicle:
            plate_display = ' '.join(part for part in [vehicle.plate_number or '', vehicle.plate_state or ''] if part).strip()

        sheet.cell(row=row_index, column=1, value=driver.full_name if driver else '')
        sheet.cell(row=row_index, column=2, value=license_display)
        sheet.cell(row=row_index, column=3, value=company.name if company else '')
        sheet.cell(row=row_index, column=4, value=phone_number)
        sheet.cell(row=row_index, column=5, value=vehicle_type)
        sheet.cell(row=row_index, column=6, value=plate_display)
        sheet.cell(row=row_index, column=7, value=vehicle.make_model_color if vehicle else '')
        sheet.cell(row=row_index, column=8, value=destination)
        sheet.cell(row=row_index, column=9, value=log.inspection_type or (driver.visit_type if driver else ''))


def _write_details_sheet(workbook, logs, log_date):
    if 'LOG DETAILS' in workbook.sheetnames:
        old_sheet = workbook['LOG DETAILS']
        workbook.remove(old_sheet)
    sheet = workbook.create_sheet('LOG DETAILS')
    sheet.append(['LOG DATE', log_date])
    sheet.append([])
    sheet.append(['ENTRY TIME UTC', 'ENTRY TIME', 'DRIVER', 'COMPANY', 'VEHICLE', 'INSPECTION TYPE', 'SCAN TOKEN', 'NOTES'])
    for log in logs:
        driver = log.driver
        company = log.company or (driver.company if driver else None)
        vehicle = log.vehicle
        plate_display = ''
        if vehicle:
            plate_display = ' '.join(part for part in [vehicle.plate_number or '', vehicle.plate_state or ''] if part).strip()
        sheet.append(
            [
                log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
                log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '',
                driver.full_name if driver else '',
                company.name if company else '',
                plate_display,
                log.inspection_type or '',
                log.scan_token or '',
                log.notes or '',
            ]
        )


def build_daily_workbook(log_date):
    if Workbook is None or load_workbook is None:
        raise RuntimeError('Truck Gate Excel export requires openpyxl to be installed.')
    logs = (
        TruckGateLog.query.filter_by(log_date=log_date)
        .order_by(TruckGateLog.created_at.asc(), TruckGateLog.id.asc())
        .all()
    )
    ensure_daily_workbook_directory(log_date)
    path = daily_workbook_abspath(log_date)
    if DEFAULT_TRUCK_GATE_SOURCE and os.path.exists(DEFAULT_TRUCK_GATE_SOURCE):
        shutil.copyfile(DEFAULT_TRUCK_GATE_SOURCE, path)
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
    _write_database_sheet(workbook, logs)
    _write_details_sheet(workbook, logs, log_date)
    workbook.save(path)
    workbook.close()
    return path
